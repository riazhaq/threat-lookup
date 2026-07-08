#!/usr/bin/env python3
import argparse
import asyncio
import aiohttp
import csv
import io
import ipaddress
import json
import os
import random
import re
import sys
import time
import urllib.error
import urllib.request
from pathlib import Path
from typing import Dict, Any, List
from urllib.parse import urlparse, quote

# Load key-value pairs from .env into process environment if present.
def load_local_env(env_path: str = ".env", override: bool = True) -> None:
    if not os.path.exists(env_path):
        return
    try:
        with open(env_path, "r", encoding="utf-8") as f:
            for raw_line in f:
                line = raw_line.strip()
                if not line or line.startswith("#") or "=" not in line:
                    continue
                key, value = line.split("=", 1)
                key = key.strip()
                value = value.strip().strip('"').strip("'")
                if key and (override or key not in os.environ):
                    os.environ[key] = value
    except OSError:
        pass


load_local_env(override=False)

# ----------------------------
# CONFIG
# ----------------------------
VT_API_KEY = os.getenv("VT_API_KEY")
OTX_API_KEY = os.getenv("OTX_API_KEY")
ABUSE_API_KEY = os.getenv("ABUSE_API_KEY")
THREATFOX_API_KEY = os.getenv("THREATFOX_API_KEY")
SILENTPUSH_API_KEY = os.getenv("SILENTPUSH_API_KEY")
SILENTPUSH_ACCESS_KEY = (os.getenv("SILENTPUSH_ACCESS_KEY", "") or SILENTPUSH_API_KEY or "").strip()
SILENTPUSH_DATA_SOURCE = os.getenv("SILENTPUSH_DATA_SOURCE", "iofa")
SILENTPUSH_API_BASE_URL = os.getenv("SILENTPUSH_API_BASE_URL", "https://api.silentpush.com/api/v1/merge-api").strip().rstrip("/")
SILENTPUSH_META_API_BASE_URL = os.getenv("SILENTPUSH_META_API_BASE_URL", "https://api.silentpush.com/api/v1/meta-api").strip().rstrip("/")
SILENTPUSH_V2_API_BASE_URL = os.getenv("SILENTPUSH_V2_API_BASE_URL", "https://api.silentpush.com/api/v2").strip().rstrip("/")
SILENTPUSH_MAX_ENRICH_CALLS = max(0, int(os.getenv("SILENTPUSH_MAX_ENRICH_CALLS", "2") or "2"))
SILENTPUSH_MAX_NON_EXPLORE_CALLS = max(0, int(os.getenv("SILENTPUSH_MAX_NON_EXPLORE_CALLS", "8") or "8"))
SILENTPUSH_ENABLE_CONTEXT_GRAPH_SEARCH = os.getenv("SILENTPUSH_ENABLE_CONTEXT_GRAPH_SEARCH", "true").strip().lower() in {"1", "true", "yes", "on"}
SILENTPUSH_ENABLE_ENRICH_INDICATOR = os.getenv("SILENTPUSH_ENABLE_ENRICH_INDICATOR", "true").strip().lower() in {"1", "true", "yes", "on"}
SILENTPUSH_ENABLE_LIVE_SCAN = os.getenv("SILENTPUSH_ENABLE_LIVE_SCAN", "true").strip().lower() in {"1", "true", "yes", "on"}
SILENTPUSH_ENABLE_THREAT_RANKING = os.getenv("SILENTPUSH_ENABLE_THREAT_RANKING", "true").strip().lower() in {"1", "true", "yes", "on"}
SILENTPUSH_ENABLE_PADNS_LOOKUP = os.getenv("SILENTPUSH_ENABLE_PADNS_LOOKUP", "true").strip().lower() in {"1", "true", "yes", "on"}
SILENTPUSH_ENABLE_BULK_ENRICH = os.getenv("SILENTPUSH_ENABLE_BULK_ENRICH", "true").strip().lower() in {"1", "true", "yes", "on"}
SILENTPUSH_ENABLE_THREAT_CHECK = os.getenv("SILENTPUSH_ENABLE_THREAT_CHECK", "true").strip().lower() in {"1", "true", "yes", "on"}
SILENTPUSH_THREAT_CHECK_BASE_URL = os.getenv("SILENTPUSH_THREAT_CHECK_BASE_URL", "https://api.threatcheck.silentpush.com").strip().rstrip("/")
SILENTPUSH_BPH_SUSPECTED_DOMAIN_COUNT = max(1, int(os.getenv("SILENTPUSH_BPH_SUSPECTED_DOMAIN_COUNT", "100") or "100"))
SILENTPUSH_BPH_LIKELY_DOMAIN_COUNT = max(1, int(os.getenv("SILENTPUSH_BPH_LIKELY_DOMAIN_COUNT", "1000") or "1000"))
SILENTPUSH_VERBOSE_NON_EXPLORE_ERRORS = os.getenv("SILENTPUSH_VERBOSE_NON_EXPLORE_ERRORS", "false").strip().lower() in {"1", "true", "yes", "on"}
SILENTPUSH_VERBOSE_EXPLORE_ERRORS = os.getenv("SILENTPUSH_VERBOSE_EXPLORE_ERRORS", "false").strip().lower() in {"1", "true", "yes", "on"}
SILENTPUSH_SPQL_PAYLOAD_MODE = os.getenv("SILENTPUSH_SPQL_PAYLOAD_MODE", "auto").strip().lower()
SILENTPUSH_SPQL_QUERY_TEMPLATE = os.getenv(
    "SILENTPUSH_SPQL_QUERY_TEMPLATE",
    "SELECT * FROM scandata WHERE query='{indicator}' LIMIT 25",
)
SILENTPUSH_SPQL_QUERY_TEMPLATE_ALT = os.getenv(
    "SILENTPUSH_SPQL_QUERY_TEMPLATE_ALT",
    "SELECT * FROM scandata WHERE indicator='{indicator}' LIMIT 25",
)
SPUR_API_KEY = os.getenv("SPUR_API_KEY")
SPUR_CONTEXT_DT = os.getenv("SPUR_CONTEXT_DT", "").strip()
SPUR_USE_MAXMIND_GEO = os.getenv("SPUR_USE_MAXMIND_GEO", "false").strip().lower() in {"1", "true", "yes", "on"}
SPUR_ENABLE_TAG_METADATA = os.getenv("SPUR_ENABLE_TAG_METADATA", "true").strip().lower() in {"1", "true", "yes", "on"}
SPUR_MAX_TAG_METADATA = max(0, int(os.getenv("SPUR_MAX_TAG_METADATA", "3") or "3"))
RL_SPECTRA_BASE_URL = os.getenv("RL_SPECTRA_BASE_URL", "").strip().rstrip("/")
RL_SPECTRA_TOKEN = os.getenv("RL_SPECTRA_TOKEN")

_SPUR_TAG_METADATA_CACHE: Dict[str, Dict[str, Any]] = {}


def refresh_runtime_config(env_path: str = ".env") -> None:
    """Reload .env and refresh module-level config used by API clients and UI."""
    global VT_API_KEY, OTX_API_KEY, ABUSE_API_KEY, THREATFOX_API_KEY
    global SILENTPUSH_API_KEY, SILENTPUSH_ACCESS_KEY, SILENTPUSH_DATA_SOURCE, SILENTPUSH_API_BASE_URL
    global SILENTPUSH_META_API_BASE_URL, SILENTPUSH_V2_API_BASE_URL
    global SILENTPUSH_MAX_ENRICH_CALLS
    global SILENTPUSH_MAX_NON_EXPLORE_CALLS
    global SILENTPUSH_ENABLE_CONTEXT_GRAPH_SEARCH, SILENTPUSH_ENABLE_ENRICH_INDICATOR
    global SILENTPUSH_ENABLE_LIVE_SCAN, SILENTPUSH_ENABLE_THREAT_RANKING
    global SILENTPUSH_ENABLE_PADNS_LOOKUP, SILENTPUSH_ENABLE_BULK_ENRICH
    global SILENTPUSH_ENABLE_THREAT_CHECK, SILENTPUSH_THREAT_CHECK_BASE_URL
    global SILENTPUSH_BPH_SUSPECTED_DOMAIN_COUNT, SILENTPUSH_BPH_LIKELY_DOMAIN_COUNT
    global SILENTPUSH_VERBOSE_NON_EXPLORE_ERRORS, SILENTPUSH_VERBOSE_EXPLORE_ERRORS, SPUR_API_KEY
    global SPUR_CONTEXT_DT, SPUR_USE_MAXMIND_GEO, SPUR_ENABLE_TAG_METADATA, SPUR_MAX_TAG_METADATA
    global SILENTPUSH_SPQL_PAYLOAD_MODE, SILENTPUSH_SPQL_QUERY_TEMPLATE, SILENTPUSH_SPQL_QUERY_TEMPLATE_ALT
    global RL_SPECTRA_BASE_URL, RL_SPECTRA_TOKEN

    load_local_env(env_path=env_path, override=True)

    VT_API_KEY = os.getenv("VT_API_KEY")
    OTX_API_KEY = os.getenv("OTX_API_KEY")
    ABUSE_API_KEY = os.getenv("ABUSE_API_KEY")
    THREATFOX_API_KEY = os.getenv("THREATFOX_API_KEY")
    SILENTPUSH_API_KEY = os.getenv("SILENTPUSH_API_KEY")
    SILENTPUSH_ACCESS_KEY = (os.getenv("SILENTPUSH_ACCESS_KEY", "") or SILENTPUSH_API_KEY or "").strip()
    SILENTPUSH_DATA_SOURCE = os.getenv("SILENTPUSH_DATA_SOURCE", "iofa")
    SILENTPUSH_API_BASE_URL = os.getenv("SILENTPUSH_API_BASE_URL", "https://api.silentpush.com/api/v1/merge-api").strip().rstrip("/")
    SILENTPUSH_META_API_BASE_URL = os.getenv("SILENTPUSH_META_API_BASE_URL", "https://api.silentpush.com/api/v1/meta-api").strip().rstrip("/")
    SILENTPUSH_V2_API_BASE_URL = os.getenv("SILENTPUSH_V2_API_BASE_URL", "https://api.silentpush.com/api/v2").strip().rstrip("/")
    SILENTPUSH_MAX_ENRICH_CALLS = max(0, int(os.getenv("SILENTPUSH_MAX_ENRICH_CALLS", "2") or "2"))
    SILENTPUSH_MAX_NON_EXPLORE_CALLS = max(0, int(os.getenv("SILENTPUSH_MAX_NON_EXPLORE_CALLS", "8") or "8"))
    SILENTPUSH_ENABLE_CONTEXT_GRAPH_SEARCH = os.getenv("SILENTPUSH_ENABLE_CONTEXT_GRAPH_SEARCH", "true").strip().lower() in {"1", "true", "yes", "on"}
    SILENTPUSH_ENABLE_ENRICH_INDICATOR = os.getenv("SILENTPUSH_ENABLE_ENRICH_INDICATOR", "true").strip().lower() in {"1", "true", "yes", "on"}
    SILENTPUSH_ENABLE_LIVE_SCAN = os.getenv("SILENTPUSH_ENABLE_LIVE_SCAN", "true").strip().lower() in {"1", "true", "yes", "on"}
    SILENTPUSH_ENABLE_THREAT_RANKING = os.getenv("SILENTPUSH_ENABLE_THREAT_RANKING", "true").strip().lower() in {"1", "true", "yes", "on"}
    SILENTPUSH_ENABLE_PADNS_LOOKUP = os.getenv("SILENTPUSH_ENABLE_PADNS_LOOKUP", "true").strip().lower() in {"1", "true", "yes", "on"}
    SILENTPUSH_ENABLE_BULK_ENRICH = os.getenv("SILENTPUSH_ENABLE_BULK_ENRICH", "true").strip().lower() in {"1", "true", "yes", "on"}
    SILENTPUSH_ENABLE_THREAT_CHECK = os.getenv("SILENTPUSH_ENABLE_THREAT_CHECK", "true").strip().lower() in {"1", "true", "yes", "on"}
    SILENTPUSH_THREAT_CHECK_BASE_URL = os.getenv("SILENTPUSH_THREAT_CHECK_BASE_URL", "https://api.threatcheck.silentpush.com").strip().rstrip("/")
    SILENTPUSH_BPH_SUSPECTED_DOMAIN_COUNT = max(1, int(os.getenv("SILENTPUSH_BPH_SUSPECTED_DOMAIN_COUNT", "100") or "100"))
    SILENTPUSH_BPH_LIKELY_DOMAIN_COUNT = max(1, int(os.getenv("SILENTPUSH_BPH_LIKELY_DOMAIN_COUNT", "1000") or "1000"))
    SILENTPUSH_VERBOSE_NON_EXPLORE_ERRORS = os.getenv("SILENTPUSH_VERBOSE_NON_EXPLORE_ERRORS", "false").strip().lower() in {"1", "true", "yes", "on"}
    SILENTPUSH_VERBOSE_EXPLORE_ERRORS = os.getenv("SILENTPUSH_VERBOSE_EXPLORE_ERRORS", "false").strip().lower() in {"1", "true", "yes", "on"}
    SILENTPUSH_SPQL_PAYLOAD_MODE = os.getenv("SILENTPUSH_SPQL_PAYLOAD_MODE", "auto").strip().lower()
    SILENTPUSH_SPQL_QUERY_TEMPLATE = os.getenv(
        "SILENTPUSH_SPQL_QUERY_TEMPLATE",
        "SELECT * FROM scandata WHERE query='{indicator}' LIMIT 25",
    )
    SILENTPUSH_SPQL_QUERY_TEMPLATE_ALT = os.getenv(
        "SILENTPUSH_SPQL_QUERY_TEMPLATE_ALT",
        "SELECT * FROM scandata WHERE indicator='{indicator}' LIMIT 25",
    )
    SPUR_API_KEY = os.getenv("SPUR_API_KEY")
    SPUR_CONTEXT_DT = os.getenv("SPUR_CONTEXT_DT", "").strip()
    SPUR_USE_MAXMIND_GEO = os.getenv("SPUR_USE_MAXMIND_GEO", "false").strip().lower() in {"1", "true", "yes", "on"}
    SPUR_ENABLE_TAG_METADATA = os.getenv("SPUR_ENABLE_TAG_METADATA", "true").strip().lower() in {"1", "true", "yes", "on"}
    SPUR_MAX_TAG_METADATA = max(0, int(os.getenv("SPUR_MAX_TAG_METADATA", "3") or "3"))
    RL_SPECTRA_BASE_URL = os.getenv("RL_SPECTRA_BASE_URL", "").strip().rstrip("/")
    RL_SPECTRA_TOKEN = os.getenv("RL_SPECTRA_TOKEN")

    _SPUR_TAG_METADATA_CACHE.clear()

TIMEOUT = 5
MAX_RETRIES = 0
RETRY_BASE_DELAY = 0.6
RETRY_JITTER = 0.25
RETRY_MAX_DELAY = 1.5

ALLOWLIST = {"8.8.8.8", "1.1.1.1", "google.com"}

_SHA256_RE = re.compile(r'^[0-9a-fA-F]{64}$')
_SHA1_RE   = re.compile(r'^[0-9a-fA-F]{40}$')
_MD5_RE    = re.compile(r'^[0-9a-fA-F]{32}$')
_URL_RE    = re.compile(r'^https?://', re.IGNORECASE)


# ----------------------------
# KEY WARNINGS
# ----------------------------
def check_keys() -> None:
    missing = []
    if not VT_API_KEY:
        missing.append("VT_API_KEY (VirusTotal)")
    if not OTX_API_KEY:
        missing.append("OTX_API_KEY (AlienVault OTX)")
    if not ABUSE_API_KEY:
        missing.append("ABUSE_API_KEY (AbuseIPDB)")
    if not THREATFOX_API_KEY:
        missing.append("THREATFOX_API_KEY (ThreatFox)")
    if not SILENTPUSH_API_KEY:
        missing.append("SILENTPUSH_API_KEY (Silent Push Explore API)")
    if not SPUR_API_KEY:
        missing.append("SPUR_API_KEY (Spur Context API)")
    if not RL_SPECTRA_BASE_URL:
        missing.append("RL_SPECTRA_BASE_URL (ReversingLabs Spectra Analyze)")
    if not RL_SPECTRA_TOKEN:
        missing.append("RL_SPECTRA_TOKEN (ReversingLabs Spectra Analyze)")
    if missing:
        print(f"[WARNING] Missing API keys — these sources will return no data: {', '.join(missing)}", file=sys.stderr)


# ----------------------------
# IOC TYPE DETECTION
# ----------------------------
def detect_type(ioc: str) -> str:
    try:
        ipaddress.ip_address(ioc)
        return "ip"
    except ValueError:
        pass
    if _SHA256_RE.match(ioc) or _SHA1_RE.match(ioc) or _MD5_RE.match(ioc):
        return "hash"
    if _URL_RE.match(ioc) or "/" in ioc:
        return "url"
    return "domain"


# ----------------------------
# API CALLS
# ----------------------------
def _retry_delay(attempt: int) -> float:
    delay = (RETRY_BASE_DELAY * (2 ** attempt)) + random.uniform(0, RETRY_JITTER)
    return min(delay, RETRY_MAX_DELAY)


def _header_value(headers: Dict[str, Any], *keys: str) -> Any:
    if not isinstance(headers, dict):
        return None
    header_map = {str(k).lower(): v for k, v in headers.items()}
    for key in keys:
        val = header_map.get(str(key).lower())
        if val not in (None, "", []):
            return val
    return None


def _to_int_or_none(value: Any):
    if value in (None, "", []):
        return None
    try:
        return int(str(value).strip())
    except Exception:
        return None


def _extract_api_usage(headers: Dict[str, Any]) -> Dict[str, Any]:
    if not isinstance(headers, dict) or not headers:
        return {}

    usage = {
        "limit": _header_value(headers, "x-ratelimit-limit", "x-rate-limit-limit", "ratelimit-limit", "x-daily-limit"),
        "remaining": _header_value(headers, "x-ratelimit-remaining", "x-rate-limit-remaining", "ratelimit-remaining", "x-balance-remaining", "x-daily-remaining"),
        "used": _header_value(headers, "x-ratelimit-used", "x-rate-limit-used", "ratelimit-used", "x-balance-used", "x-daily-used"),
        "reset": _header_value(headers, "x-ratelimit-reset", "x-rate-limit-reset", "ratelimit-reset", "x-ratelimit-reset-after", "retry-after"),
        "window": _header_value(headers, "x-ratelimit-window", "x-rate-limit-window", "ratelimit-window"),
    }

    limit_i = _to_int_or_none(usage.get("limit"))
    remaining_i = _to_int_or_none(usage.get("remaining"))
    used_i = _to_int_or_none(usage.get("used"))
    if used_i is None and limit_i is not None and remaining_i is not None:
        usage["used"] = max(0, limit_i - remaining_i)

    return {k: v for k, v in usage.items() if v not in (None, "", [])}


async def _request_json_with_retry(session, method: str, url: str, return_headers: bool = False, **kwargs):
    for attempt in range(MAX_RETRIES + 1):
        try:
            async with session.request(method, url, timeout=TIMEOUT, **kwargs) as resp:
                if resp.status == 429 and attempt < MAX_RETRIES:
                    await asyncio.sleep(_retry_delay(attempt))
                    continue
                try:
                    data = await resp.json()
                except Exception:
                    data = {}
                if return_headers:
                    return resp.status, data, None, dict(resp.headers)
                return resp.status, data, None
        except asyncio.TimeoutError:
            if attempt < MAX_RETRIES:
                await asyncio.sleep(_retry_delay(attempt))
                continue
            if return_headers:
                return None, {}, "timeout", {}
            return None, {}, "timeout"
        except Exception as e:
            if return_headers:
                return None, {}, str(e), {}
            return None, {}, str(e)

    if return_headers:
        return None, {}, "timeout", {}
    return None, {}, "timeout"


async def query_virustotal(session, ioc, ioc_type):
    if not VT_API_KEY:
        return {"source": "virustotal", "malicious": 0, "suspicious": 0, "score": 0, "error": "no_api_key"}

    url_map = {
        "ip":     f"https://www.virustotal.com/api/v3/ip_addresses/{ioc}",
        "domain": f"https://www.virustotal.com/api/v3/domains/{ioc}",
        "url":    f"https://www.virustotal.com/api/v3/urls/{ioc}",
        "hash":   f"https://www.virustotal.com/api/v3/files/{ioc}",
    }
    headers = {"x-apikey": VT_API_KEY}

    status, data, request_error = await _request_json_with_retry(
        session,
        "GET",
        url_map[ioc_type],
        headers=headers,
    )
    if request_error:
        return {"source": "virustotal", "malicious": 0, "suspicious": 0, "score": 0, "error": request_error}
    if status == 401:
        return {"source": "virustotal", "malicious": 0, "suspicious": 0, "score": 0, "error": "invalid_api_key"}
    if status == 429:
        return {"source": "virustotal", "malicious": 0, "suspicious": 0, "score": 0, "error": "rate_limited"}
    if status == 404:
        return {"source": "virustotal", "malicious": 0, "suspicious": 0, "score": 0, "error": "not_found"}

    attrs = data.get("data", {}).get("attributes", {})
    stats = attrs.get("last_analysis_stats", {})
    malicious  = stats.get("malicious", 0)
    suspicious = stats.get("suspicious", 0)

    # Context extraction
    context = {}
    # Top malicious engine names (up to 5)
    analysis_results = attrs.get("last_analysis_results", {})
    flagging_engines = [
        eng for eng, res in analysis_results.items()
        if res.get("category") == "malicious"
    ][:5]
    if flagging_engines:
        context["detected_by"] = flagging_engines
    # IP-specific
    if ioc_type == "ip":
        for field in ("country", "continent", "as_owner", "asn", "network"):
            val = attrs.get(field)
            if val is not None:
                context[field] = val
    # Domain-specific
    if ioc_type == "domain":
        categories = attrs.get("categories", {})
        if categories:
            context["categories"] = list(categories.values())
        for field in ("creation_date", "registrar", "country"):
            val = attrs.get(field)
            if val is not None:
                context[field] = val
    # Hash-specific
    if ioc_type == "hash":
        for field in ("meaningful_name", "type_description", "magic"):
            val = attrs.get(field)
            if val is not None:
                context[field] = val
        names = attrs.get("names", [])
        if names:
            context["known_filenames"] = names[:5]
    # Data freshness and community reputation (all types)
    last_analysis_date = attrs.get("last_analysis_date")
    if last_analysis_date:
        import datetime
        context["last_analysed"] = datetime.datetime.fromtimestamp(last_analysis_date, tz=datetime.timezone.utc).strftime("%Y-%m-%d %H:%M UTC")
    reputation = attrs.get("reputation")
    if reputation is not None:
        context["community_reputation"] = reputation
    votes = attrs.get("total_votes", {})
    if votes:
        context["community_votes"] = votes

    # URL-specific
    if ioc_type == "url":
        for field, key in (("final_url", "last_final_url"), ("http_status", "last_http_response_code"), ("page_title", "title")):
            val = attrs.get(key)
            if val is not None:
                context[field] = val

    # Hash-specific extras
    if ioc_type == "hash":
        for field, key in (("first_submitted", "first_submission_date"), ("times_submitted", "times_submitted")):
            val = attrs.get(key)
            if val is not None:
                if key == "first_submission_date":
                    import datetime
                    context[field] = datetime.datetime.fromtimestamp(val, tz=datetime.timezone.utc).strftime("%Y-%m-%d")
                else:
                    context[field] = val
        sandbox_verdicts = attrs.get("sandbox_verdicts", {})
        if sandbox_verdicts:
            context["sandbox_verdicts"] = {
                sb: v.get("category") for sb, v in list(sandbox_verdicts.items())[:5]
            }

    # Tags (all types)
    tags = attrs.get("tags", [])
    if tags:
        context["tags"] = tags[:10]

    return {
        "source": "virustotal",
        "malicious": malicious,
        "suspicious": suspicious,
        "score": malicious,
        **context,
    }


async def query_alienvault(session, ioc, ioc_type):
    if not OTX_API_KEY:
        return {"source": "alienvault", "pulses": 0, "score": 0, "error": "no_api_key"}

    otx_type_map = {"ip": "IPv4", "domain": "domain", "url": "url", "hash": "file"}
    otx_type = otx_type_map.get(ioc_type, "IPv4")
    url = f"https://otx.alienvault.com/api/v1/indicators/{otx_type}/{ioc}/general"
    headers = {"X-OTX-API-KEY": OTX_API_KEY}

    status, data, request_error = await _request_json_with_retry(
        session,
        "GET",
        url,
        headers=headers,
    )
    if request_error:
        return {"source": "alienvault", "pulses": 0, "score": 0, "error": request_error}
    if status == 401:
        return {"source": "alienvault", "pulses": 0, "score": 0, "error": "invalid_api_key"}
    if status == 429:
        return {"source": "alienvault", "pulses": 0, "score": 0, "error": "rate_limited"}

    pulse_info = data.get("pulse_info", {})
    pulse_list = pulse_info.get("pulses", [])
    pulses = len(pulse_list)

    # Context extraction
    context = {}
    # Geo/network for IPs
    for field in ("country_name", "city", "asn"):
        val = data.get(field)
        if val:
            context[field] = val
    # Pulse names (up to 5)
    pulse_names = [p.get("name") for p in pulse_list[:5] if p.get("name")]
    if pulse_names:
        context["pulse_names"] = pulse_names
    # Aggregate malware families across all pulses
    malware_families = list({
        mf.get("display_name") or mf.get("id")
        for p in pulse_list
        for mf in p.get("malware_families", [])
        if mf.get("display_name") or mf.get("id")
    })
    if malware_families:
        context["malware_families"] = malware_families[:10]
    # Aggregate tags across all pulses
    all_tags = list({
        tag
        for p in pulse_list
        for tag in p.get("tags", [])
        if tag
    })
    if all_tags:
        context["tags"] = all_tags[:15]
    # Targeted countries
    targeted = list({
        c
        for p in pulse_list
        for c in p.get("targeted_countries", [])
        if c
    })
    if targeted:
        context["targeted_countries"] = targeted[:10]
    # MITRE ATT&CK IDs
    attack_ids = list({
        a.get("display_name") or a.get("id")
        for p in pulse_list
        for a in p.get("attack_ids", [])
        if a.get("display_name") or a.get("id")
    })
    if attack_ids:
        context["attack_ids"] = attack_ids[:10]
    # Adversary / threat actor names
    adversaries = list({
        p.get("adversary")
        for p in pulse_list
        if p.get("adversary")
    })
    if adversaries:
        context["adversaries"] = adversaries[:5]
    # Industries targeted
    industries = list({
        ind
        for p in pulse_list
        for ind in p.get("industries", [])
        if ind
    })
    if industries:
        context["industries_targeted"] = industries[:10]

    return {
        "source": "alienvault",
        "pulses": pulses,
        "score": pulses,
        **context,
    }


async def query_abuseipdb(session, ioc):
    if not ABUSE_API_KEY:
        return {"source": "abuseipdb", "abuse_score": 0, "total_reports": 0, "score": 0, "error": "no_api_key"}

    url = "https://api.abuseipdb.com/api/v2/check"
    headers = {"Key": ABUSE_API_KEY, "Accept": "application/json"}
    params = {"ipAddress": ioc, "maxAgeInDays": 90}

    status, data, request_error = await _request_json_with_retry(
        session,
        "GET",
        url,
        headers=headers,
        params=params,
    )
    if request_error:
        return {"source": "abuseipdb", "abuse_score": 0, "total_reports": 0, "score": 0, "error": request_error}
    if status == 401:
        return {"source": "abuseipdb", "abuse_score": 0, "total_reports": 0, "score": 0, "error": "invalid_api_key"}
    if status == 429:
        return {"source": "abuseipdb", "abuse_score": 0, "total_reports": 0, "score": 0, "error": "rate_limited"}

    d = data.get("data", {})
    abuse_score   = d.get("abuseConfidenceScore", 0)
    total_reports = d.get("totalReports", 0)

    context = {}
    for field, key in (
        ("country",          "countryCode"),
        ("isp",              "isp"),
        ("domain",           "domain"),
        ("usage_type",       "usageType"),
        ("is_tor",           "isTor"),
        ("is_whitelisted",   "isWhitelisted"),
        ("last_reported_at", "lastReportedAt"),
    ):
        val = d.get(key)
        if val is not None:
            context[field] = val
    distinct_reporters = d.get("numDistinctUsers")
    if distinct_reporters is not None:
        context["distinct_reporters"] = distinct_reporters
    hostnames = d.get("hostnames", [])
    if hostnames:
        context["hostnames"] = hostnames[:5]

    return {
        "source": "abuseipdb",
        "abuse_score": abuse_score,
        "total_reports": total_reports,
        "score": abuse_score,
        **context,
    }


async def query_threatfox(session, ioc, ioc_type):
    if not THREATFOX_API_KEY:
        return {"source": "threatfox", "matches": 0, "max_confidence": 0, "score": 0, "error": "no_api_key"}

    url = "https://threatfox-api.abuse.ch/api/v1/"
    headers = {"Auth-Key": THREATFOX_API_KEY, "Content-Type": "application/json"}
    parsed_url = urlparse(ioc) if ioc_type == "url" else None
    candidates: List[str] = [ioc]
    if ioc_type == "url" and parsed_url:
        host = (parsed_url.hostname or "").strip()
        if host:
            candidates.append(host)
        path_root = ioc.split("?", 1)[0].rstrip("/")
        if path_root and path_root != ioc:
            candidates.append(path_root)
    elif ioc_type == "domain":
        domain_norm = ioc.strip().rstrip(".").lower()
        if domain_norm and domain_norm != ioc:
            candidates.append(domain_norm)

    deduped_candidates: List[str] = []
    seen_candidates = set()
    for candidate in candidates:
        key = candidate.strip().lower()
        if not key or key in seen_candidates:
            continue
        seen_candidates.add(key)
        deduped_candidates.append(candidate)

    matches: List[Dict[str, Any]] = []
    statuses: List[str] = []
    for candidate in deduped_candidates:
        for exact in (True, False):
            payload = {"query": "search_ioc", "search_term": candidate, "exact_match": exact}
            status, data, request_error = await _request_json_with_retry(
                session,
                "POST",
                url,
                headers=headers,
                json=payload,
            )

            if request_error:
                return {"source": "threatfox", "matches": 0, "max_confidence": 0, "score": 0, "error": request_error}
            if status == 401:
                return {"source": "threatfox", "matches": 0, "max_confidence": 0, "score": 0, "error": "invalid_api_key"}
            if status == 429:
                return {"source": "threatfox", "matches": 0, "max_confidence": 0, "score": 0, "error": "rate_limited"}

            query_status = str(data.get("query_status") or "").lower()
            if query_status == "ok":
                candidate_matches = data.get("data", []) or []
                if candidate_matches:
                    matches = candidate_matches
                    break

            if query_status:
                statuses.append(f"{candidate}|exact={str(exact).lower()}:{query_status}")
            else:
                statuses.append(f"{candidate}|exact={str(exact).lower()}:unknown")

            # Do not run non-exact immediately after an exact error for hash IOCs.
            if ioc_type == "hash" and exact:
                break
        if matches:
            break

    if not matches:
        out = {"source": "threatfox", "matches": 0, "max_confidence": 0, "score": 0, "error": "no_result"}
        if statuses:
            out["attempts"] = statuses[:6]
        return out

    max_confidence = max((match.get("confidence_level", 0) or 0) for match in matches) if matches else 0
    score = round(min(max_confidence / 5, 20)) if matches else 0

    # Context extraction across all matches
    malware_families = list({
        m.get("malware_printable") or m.get("malware")
        for m in matches
        if m.get("malware_printable") or m.get("malware")
    })
    threat_types = list({
        m.get("threat_type_desc") or m.get("threat_type")
        for m in matches
        if m.get("threat_type_desc") or m.get("threat_type")
    })
    all_tags = list({
        tag
        for m in matches
        for tag in (m.get("tags") or [])
        if tag
    })
    first_seen_dates = sorted({
        m.get("first_seen")
        for m in matches
        if m.get("first_seen")
    })
    references = list({
        m.get("reference")
        for m in matches
        if m.get("reference")
    })

    result = {
        "source": "threatfox",
        "matches": len(matches),
        "ioc_types": sorted({match.get("ioc_type", "") for match in matches if match.get("ioc_type")}),
        "max_confidence": max_confidence,
        "score": score,
    }
    if malware_families:
        result["malware_families"] = malware_families
    if threat_types:
        result["threat_types"] = threat_types
    if all_tags:
        result["tags"] = all_tags[:10]
    if first_seen_dates:
        result["first_seen"] = first_seen_dates[0]
    if references:
        result["references"] = references[:3]
    # Last seen and malware aliases
    last_seen_dates = sorted({
        m.get("last_seen")
        for m in matches
        if m.get("last_seen")
    }, reverse=True)
    if last_seen_dates:
        result["last_seen"] = last_seen_dates[0]
    malware_aliases = list({
        alias
        for m in matches
        for alias in (m.get("malware_alias") or "").split(",")
        if alias.strip()
    })
    if malware_aliases:
        result["malware_aliases"] = malware_aliases[:10]
    return result


def _silentpush_indicator_for_ioc(ioc: str, ioc_type: str):
    if ioc_type == "ip":
        return "ip", ioc
    if ioc_type == "domain":
        return "domain", ioc
    if ioc_type == "url":
        parsed = urlparse(ioc)
        host = (parsed.hostname or "").strip()
        if host:
            return "domain", host
        return None, None
    return None, None


def _silentpush_score_from_risk(risk_score: int) -> int:
    bounded = max(0, min(100, int(risk_score or 0)))
    return min(20, round(bounded / 5))


def _silentpush_is_listed_from_risk(risk_score: int) -> bool:
    return int(risk_score or 0) >= 80


def _silentpush_verdict_from_risk(risk_score: int) -> str:
    bounded = max(0, min(100, int(risk_score or 0)))
    if bounded >= 80:
        return "Malicious"
    if bounded >= 50:
        return "Suspicious"
    if bounded > 0:
        return "Low Risk"
    return "Benign"


def _spur_verdict_from_signals(risks: List[str], tunnel_count: int, proxy_count: int) -> str:
    if risks:
        return "Malicious"
    if tunnel_count > 0 or proxy_count > 0:
        return "Suspicious"
    return "Benign"


def _spur_sanitized_dt() -> str:
    raw = (SPUR_CONTEXT_DT or "").strip()
    if not raw:
        return ""
    return raw if re.match(r"^\d{8}$", raw) else ""


async def _spur_fetch_tag_metadata(session, tag: str) -> Dict[str, Any]:
    tag_norm = str(tag or "").strip().upper()
    if not tag_norm:
        return {}

    cached = _SPUR_TAG_METADATA_CACHE.get(tag_norm)
    if cached is not None:
        return cached

    status, data, request_error = await _request_json_with_retry(
        session,
        "GET",
        f"https://api.spur.us/v2/metadata/tags/{quote(tag_norm, safe='')}",
        headers={"Token": SPUR_API_KEY},
    )
    if request_error or status in {400, 401, 403, 404, 429} or (status and status >= 500) or not isinstance(data, dict):
        _SPUR_TAG_METADATA_CACHE[tag_norm] = {}
        return {}

    metadata = {
        "name": data.get("name"),
        "description": data.get("description"),
        "website": data.get("website"),
        "categories": (data.get("categories") or [])[:8],
        "protocols": (data.get("protocols") or [])[:8],
        "platforms": (data.get("platforms") or [])[:8],
        "is_anonymous": data.get("isAnonymous"),
        "is_callback_proxy": data.get("isCallbackProxy"),
        "is_tracked": data.get("isTracked"),
        "max_age_days": data.get("maxAgeDays"),
        "metrics": data.get("metrics") if isinstance(data.get("metrics"), dict) else None,
    }
    metadata = {k: v for k, v in metadata.items() if v not in (None, "", [], {})}
    _SPUR_TAG_METADATA_CACHE[tag_norm] = metadata
    return metadata


def _silentpush_compact_payload(value: Any, depth: int = 0, max_depth: int = 2, max_items: int = 10):
    if depth >= max_depth:
        if isinstance(value, (dict, list)):
            return f"<{type(value).__name__}>"
        return value

    if isinstance(value, dict):
        compact = {}
        for idx, (k, v) in enumerate(value.items()):
            if idx >= max_items:
                break
            compact[k] = _silentpush_compact_payload(v, depth + 1, max_depth, max_items)
        return compact

    if isinstance(value, list):
        return [
            _silentpush_compact_payload(v, depth + 1, max_depth, max_items)
            for v in value[:max_items]
        ]

    return value


def _silentpush_iter_leaf_items(value: Any, prefix: str = "", depth: int = 0, max_depth: int = 5):
    if depth > max_depth:
        return
    if isinstance(value, dict):
        for k, v in value.items():
            key = f"{prefix}.{k}" if prefix else str(k)
            yield from _silentpush_iter_leaf_items(v, key, depth + 1, max_depth)
        return
    if isinstance(value, list):
        for idx, item in enumerate(value[:15]):
            key = f"{prefix}[{idx}]"
            yield from _silentpush_iter_leaf_items(item, key, depth + 1, max_depth)
        return
    yield prefix, value


def _silentpush_collect_values(payloads: List[Any], key_terms: tuple, limit: int = 8) -> List[str]:
    out: List[str] = []
    seen = set()
    for payload in payloads:
        for key, value in _silentpush_iter_leaf_items(payload):
            key_l = key.lower()
            if not any(term in key_l for term in key_terms):
                continue
            if value is None:
                continue
            val_str = str(value).strip()
            if not val_str:
                continue
            canonical = val_str.lower()
            if canonical in seen:
                continue
            seen.add(canonical)
            out.append(val_str)
            if len(out) >= limit:
                return out
    return out


def _silentpush_normalize_asn(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).strip().upper()
    if not text:
        return ""
    if text.startswith("AS"):
        text = text[2:]
    if not text.isdigit():
        return ""
    try:
        asn_num = int(text)
    except ValueError:
        return ""
    if asn_num <= 0 or asn_num > 4294967295:
        return ""
    return str(asn_num)


def _silentpush_extract_asn_for_probe(payloads: List[Any]) -> str:
    asn_keys = {"answer_asn", "asn", "as_number", "asn_number"}
    for payload in payloads:
        for key, value in _silentpush_iter_leaf_items(payload):
            if not key:
                continue
            key_tail = key.split(".")[-1]
            if "[" in key_tail:
                key_tail = key_tail.split("[")[0]
            if key_tail not in asn_keys:
                continue
            normalized = _silentpush_normalize_asn(value)
            if normalized:
                return normalized
    return ""


def _silentpush_parse_threat_check_listed(payload: Any) -> Any:
    if isinstance(payload, bool):
        return payload
    if isinstance(payload, (int, float)):
        return bool(payload)
    if isinstance(payload, str):
        txt = payload.strip().lower()
        if txt in {"true", "1", "yes", "listed", "malicious"}:
            return True
        if txt in {"false", "0", "no", "not_listed", "benign"}:
            return False
        return None
    if isinstance(payload, dict):
        for key in ("listed", "is_listed", "match", "matched", "found", "result", "threat"):
            if key in payload:
                parsed = _silentpush_parse_threat_check_listed(payload.get(key))
                if parsed is not None:
                    return parsed
    return None


def _silentpush_extract_padns_domains(payload: Any) -> List[str]:
    if not isinstance(payload, (dict, list)):
        return []

    candidates: List[Dict[str, Any]] = []
    queue: List[Any] = [payload]
    while queue:
        node = queue.pop(0)
        if isinstance(node, list):
            for item in node:
                if isinstance(item, dict):
                    candidates.append(item)
                elif isinstance(item, list):
                    queue.append(item)
            continue
        if not isinstance(node, dict):
            continue
        for key in ("records", "results", "items", "data", "list"):
            nested = node.get(key)
            if isinstance(nested, list):
                queue.append(nested)

    domains: List[str] = []
    seen = set()
    domain_keys = ("domain", "fqdn", "rrname", "name", "hostname", "host")
    for row in candidates:
        for key in domain_keys:
            raw = row.get(key)
            if raw in (None, ""):
                continue
            value = str(raw).strip().strip(".").lower()
            if not value or " " in value or "/" in value:
                continue
            if "." not in value:
                continue
            try:
                ipaddress.ip_address(value)
                continue
            except ValueError:
                pass
            if value in seen:
                continue
            seen.add(value)
            domains.append(value)
    return domains


def _silentpush_bulletproof_hosting_signal_from_padns(payload: Any):
    if payload in (None, "", {}):
        return "insufficient_data", None, [], "PADNS reverse A data not available"

    domains = _silentpush_extract_padns_domains(payload)
    domain_count = len(domains)
    sample = domains[:10]

    if domain_count >= SILENTPUSH_BPH_LIKELY_DOMAIN_COUNT:
        return (
            "likely",
            domain_count,
            sample,
            f"PADNS reverse A observed {domain_count} domains resolving to this IP (>= {SILENTPUSH_BPH_LIKELY_DOMAIN_COUNT})",
        )
    if domain_count >= SILENTPUSH_BPH_SUSPECTED_DOMAIN_COUNT:
        return (
            "suspected",
            domain_count,
            sample,
            f"PADNS reverse A observed {domain_count} domains resolving to this IP (>= {SILENTPUSH_BPH_SUSPECTED_DOMAIN_COUNT})",
        )
    return (
        "not_evident",
        domain_count,
        sample,
        f"PADNS reverse A observed {domain_count} domains; no strong bulletproof-hosting pattern",
    )


def _silentpush_extract_primary_riskscore_context(payload: Any, query_type: str) -> Dict[str, Any]:
    context: Dict[str, Any] = {}
    if not isinstance(payload, dict):
        return context

    context["riskscore_payload"] = _silentpush_compact_payload(payload, max_depth=5, max_items=20)

    top_row = {}
    if query_type == "ip":
        ip_rows = payload.get("ip2asn") or []
        if ip_rows and isinstance(ip_rows[0], dict):
            top_row = ip_rows[0]
    else:
        top_row = payload

    if not top_row:
        return context

    context["riskscore_row"] = _silentpush_compact_payload(top_row, max_depth=5, max_items=20)

    field_aliases = {
        "ip": ("ip", "query", "indicator"),
        "answer": ("answer", "resolved_ip", "value"),
        "answer_asn": ("answer_asn", "as_number", "asn", "asn_number"),
        "answer_as_name": ("answer_as_name", "as_name", "as_org", "as_owner"),
        "country_code": ("country_code", "country"),
        "first_seen": ("first_seen", "firstseen"),
        "last_seen": ("last_seen", "lastseen"),
        "subnet": ("subnet", "cidr", "netblock", "prefix"),
        "asn_reputation": ("asn_reputation",),
        "active_ips": ("active_ips",),
        "active_subnets": ("active_subnets",),
        "ips_in_asn": ("ips_in_asn",),
        "ips_listed": ("ips_listed",),
        "registration_date": ("registration_date", "created", "created_at"),
        "last_changed_date": ("last_changed_date", "updated", "updated_at"),
        "expiration_date": ("expiration_date", "expires", "expires_at"),
        "whois_server": ("whois_server",),
        "url": ("url",),
    }
    for output_key, aliases in field_aliases.items():
        for alias in aliases:
            if top_row.get(alias) not in (None, "", []):
                context[output_key] = top_row.get(alias)
                break

    explain = top_row.get("sp_risk_score_explain")
    if isinstance(explain, dict) and explain:
        context["sp_risk_score_explain"] = _silentpush_compact_payload(explain, max_depth=5, max_items=20)
        decider = explain.get("sp_risk_score_decider")
        if decider:
            context["sp_risk_score_decider"] = decider

    return context


def _silentpush_render_template(template: str, indicator: str, ioc: str, ioc_type: str) -> str:
    raw = (template or "").strip()
    if not raw:
        return ""
    mapping = {
        "indicator": indicator,
        "ioc": ioc,
        "query": indicator,
        "type": ioc_type,
        "ioc_type": ioc_type,
    }
    try:
        return raw.format(**mapping)
    except Exception:
        return raw


def _silentpush_context_graph_variants(indicator: str, ioc: str, ioc_type: str) -> List[Dict[str, Any]]:
    mode = (SILENTPUSH_SPQL_PAYLOAD_MODE or "auto").strip().lower()
    if mode not in {"auto", "query", "spql", "statement"}:
        mode = "auto"

    keys = [mode] if mode != "auto" else ["query", "spql", "statement"]

    rendered_queries: List[str] = []
    for template in (SILENTPUSH_SPQL_QUERY_TEMPLATE, SILENTPUSH_SPQL_QUERY_TEMPLATE_ALT):
        rendered = _silentpush_render_template(template, indicator, ioc, ioc_type)
        if rendered and rendered not in rendered_queries:
            rendered_queries.append(rendered)

    if not rendered_queries:
        rendered_queries.append(f"SELECT * FROM scandata WHERE query='{indicator}' LIMIT 25")

    variants: List[Dict[str, Any]] = []
    for query in rendered_queries[:2]:
        for key in keys:
            payload = {key: query}
            variants.append(
                _silentpush_variant(
                    "POST",
                    "/explore/spql/search",
                    payload=payload,
                    base_url=SILENTPUSH_META_API_BASE_URL,
                )
            )

    return variants


def _silentpush_unwrap_payload(data: Any) -> Any:
    if isinstance(data, dict) and data.get("response") is not None:
        return data.get("response")
    return data


def _silentpush_payload_looks_not_found(payload: Any) -> bool:
    if isinstance(payload, str):
        txt = payload.strip().lower()
        if not txt:
            return False
        markers = (
            "404 page not found",
            "page not found",
            "not found",
            "no route",
            "cannot find",
        )
        return any(marker in txt for marker in markers)

    if isinstance(payload, dict):
        for key in ("error", "detail", "message", "status", "code"):
            val = payload.get(key)
            if val is None:
                continue
            text = str(val).strip().lower()
            if "not found" in text or text in {"404", "http_404", "404.0"}:
                return True
    return False


def _silentpush_variant(
    method: str,
    endpoint: str,
    params: Dict[str, Any] = None,
    payload: Dict[str, Any] = None,
    base_url: str = None,
) -> Dict[str, Any]:
    v: Dict[str, Any] = {"method": method, "endpoint": endpoint}
    if params:
        v["params"] = params
    if payload:
        v["payload"] = payload
    if base_url:
        v["base_url"] = base_url
    return v


async def _silentpush_probe_api_family(session, headers: Dict[str, str], api_name: str, variants: List[Dict[str, Any]]):
    unavailable: List[str] = []
    for variant in variants:
        method = str(variant.get("method", "GET")).upper()
        endpoint = str(variant.get("endpoint", ""))
        if not endpoint:
            continue

        kwargs: Dict[str, Any] = {"headers": headers}
        params = variant.get("params")
        payload = variant.get("payload")
        base_url = str(variant.get("base_url") or SILENTPUSH_API_BASE_URL).strip().rstrip("/")
        if isinstance(params, dict):
            kwargs["params"] = params
        if isinstance(payload, dict):
            kwargs["json"] = payload

        status, data, request_error = await _request_json_with_retry(
            session,
            method,
            f"{base_url}{endpoint}",
            **kwargs,
        )
        label = f"{api_name}:{method} {endpoint}"
        if request_error:
            unavailable.append(f"{label}:error")
            continue

        if status == 200:
            payload_unwrapped = _silentpush_unwrap_payload(data)
            if _silentpush_payload_looks_not_found(payload_unwrapped):
                unavailable.append(f"{label}:404_like")
                continue
            return {
                "api": api_name,
                "method": method,
                "endpoint": endpoint,
                "payload": payload_unwrapped,
                "payload_compact": _silentpush_compact_payload(payload_unwrapped, max_depth=5, max_items=20),
            }, unavailable

        unavailable.append(f"{label}:{status}")
        if status in {401, 403}:
            break

    return None, unavailable


async def query_silentpush(session, ioc, ioc_type):
    if not SILENTPUSH_API_KEY:
        return {"source": "silentpush", "is_listed": False, "score": 0, "error": "no_api_key"}

    query_type, indicator = _silentpush_indicator_for_ioc(ioc, ioc_type)
    if not query_type or not indicator:
        return {
            "source": "silentpush",
            "is_listed": False,
            "score": 0,
            "error": "unsupported_ioc_type",
        }

    indicator_enc = quote(indicator, safe="")
    if query_type == "ip":
        endpoint = f"/explore/ipv4/riskscore/{indicator_enc}"
    else:
        endpoint = f"/explore/domain/riskscore/{indicator_enc}"
    url = f"{SILENTPUSH_API_BASE_URL}{endpoint}"
    headers = {"X-API-KEY": SILENTPUSH_API_KEY}

    status, data, request_error, response_headers = await _request_json_with_retry(
        session,
        "GET",
        url,
        headers=headers,
        return_headers=True,
    )
    api_usage = _extract_api_usage(response_headers)
    if request_error:
        out = {"source": "silentpush", "is_listed": False, "score": 0, "error": request_error}
        if api_usage:
            out["api_usage"] = api_usage
        return out
    if status == 401:
        out = {"source": "silentpush", "is_listed": False, "score": 0, "error": "invalid_api_key"}
        if api_usage:
            out["api_usage"] = api_usage
        return out
    if status == 403:
        out = {"source": "silentpush", "is_listed": False, "score": 0, "error": "forbidden"}
        if api_usage:
            out["api_usage"] = api_usage
        return out
    if status == 429:
        out = {"source": "silentpush", "is_listed": False, "score": 0, "error": "rate_limited"}
        if api_usage:
            out["api_usage"] = api_usage
        return out
    if status == 404:
        out = {"source": "silentpush", "is_listed": False, "score": 0, "error": "not_found"}
        if api_usage:
            out["api_usage"] = api_usage
        return out
    if status in {400, 422}:
        detail = data.get("error") or data.get("detail") or "invalid_query"
        out = {
            "source": "silentpush",
            "is_listed": False,
            "score": 0,
            "error": "invalid_request" if status == 400 else "invalid_query",
            "error_detail": detail,
            "hint": "Silent Push rejected this request. Confirm SILENTPUSH_API_KEY is a valid Explore API key and the IOC format is supported.",
        }
        if api_usage:
            out["api_usage"] = api_usage
        return out

    payload = data.get("response") if isinstance(data, dict) else {}
    if not isinstance(payload, dict):
        payload = {}
    top_row: Dict[str, Any] = {}
    if query_type == "ip":
        ip_rows = payload.get("ip2asn") or []
        top_row = ip_rows[0] if ip_rows and isinstance(ip_rows[0], dict) else {}
        risk_score = int(top_row.get("sp_risk_score", 0) or 0)
    else:
        risk_score = int(payload.get("sp_risk_score", 0) or 0)

    is_listed = _silentpush_is_listed_from_risk(risk_score)
    sp_verdict = _silentpush_verdict_from_risk(risk_score)
    score = _silentpush_score_from_risk(risk_score)
    listed_txt = f"Silent Push risk score {risk_score}/100"
    context: Dict[str, Any] = _silentpush_extract_primary_riskscore_context(payload, query_type)

    # Best-effort Explore enrichment to use broader main-key capabilities
    enrich_candidates = {
        "ip": [
            f"/explore/enrich/ipv4/{indicator_enc}?explain=1",
            f"/explore/ipv4/enrichment/{indicator_enc}",
        ],
        "domain": [
            f"/explore/domain/whois/{indicator_enc}",
            f"/explore/padns/lookup/query/a/{indicator_enc}?limit=1000",
            f"/explore/enrich/domain/{indicator_enc}?explain=1",
            f"/explore/domain/whoislive/{indicator_enc}",
            f"/explore/domain/domaininfo/{indicator_enc}",
            f"/explore/domain/enrichment/{indicator_enc}",
        ],
    }

    attempted = enrich_candidates.get(query_type, [])[:SILENTPUSH_MAX_ENRICH_CALLS]
    available_endpoints: List[str] = []
    unavailable_endpoints: List[str] = []
    enrichment: Dict[str, Any] = {}
    enrichment_payloads: List[Any] = []

    for enrich_endpoint in attempted:
        enrich_url = f"{SILENTPUSH_API_BASE_URL}{enrich_endpoint}"
        enrich_status, enrich_data, enrich_error = await _request_json_with_retry(
            session,
            "GET",
            enrich_url,
            headers=headers,
        )

        if enrich_error:
            unavailable_endpoints.append(f"{enrich_endpoint}:error")
            continue

        if enrich_status == 200:
            available_endpoints.append(enrich_endpoint)
            enrich_payload = enrich_data.get("response") if isinstance(enrich_data, dict) else enrich_data
            if enrich_payload is None:
                enrich_payload = enrich_data
            if _silentpush_payload_looks_not_found(enrich_payload):
                available_endpoints.pop()
                unavailable_endpoints.append(f"{enrich_endpoint}:404_like")
                continue
            enrichment[enrich_endpoint] = _silentpush_compact_payload(enrich_payload, max_depth=5, max_items=20)
            if isinstance(enrich_payload, (dict, list)):
                enrichment_payloads.append(enrich_payload)
            continue

        unavailable_endpoints.append(f"{enrich_endpoint}:{enrich_status}")

    if available_endpoints:
        context["available_endpoints"] = available_endpoints
    if unavailable_endpoints and SILENTPUSH_VERBOSE_EXPLORE_ERRORS:
        context["unavailable_endpoints"] = unavailable_endpoints
    context["explore_enrichment_status"] = f"available_{len(available_endpoints)}_of_{len(attempted)}"
    if enrichment:
        context["enrichment"] = enrichment

    if enrichment_payloads:
        asn_vals = _silentpush_collect_values(enrichment_payloads, ("asn", "as_number", "asn_number"), limit=3)
        as_org_vals = _silentpush_collect_values(enrichment_payloads, ("as_name", "as_org", "as_owner", "organization", "org"), limit=3)
        country_vals = _silentpush_collect_values(enrichment_payloads, ("country", "country_code"), limit=3)
        registrar_vals = _silentpush_collect_values(enrichment_payloads, ("registrar",), limit=2)
        first_seen_vals = _silentpush_collect_values(enrichment_payloads, ("first_seen", "firstseen"), limit=2)
        last_seen_vals = _silentpush_collect_values(enrichment_payloads, ("last_seen", "lastseen", "updated"), limit=2)
        tag_vals = _silentpush_collect_values(enrichment_payloads, ("tag",), limit=10)
        ns_vals = _silentpush_collect_values(enrichment_payloads, ("nameserver", "name_server", "ns"), limit=6)
        dns_vals = _silentpush_collect_values(
            enrichment_payloads,
            ("a_record", "aaaa", "cname", "mx", "txt", "ptr", "dns", "resolved", "ip"),
            limit=10,
        )

        if asn_vals:
            context["asn"] = asn_vals[0]
        if as_org_vals:
            context["as_org"] = as_org_vals[0]
        if country_vals:
            context["country"] = country_vals[0]
        if registrar_vals:
            context["registrar"] = registrar_vals[0]
        if first_seen_vals:
            context["first_seen"] = first_seen_vals[0]
        if last_seen_vals:
            context["last_seen"] = last_seen_vals[0]
        if tag_vals:
            context["tags"] = tag_vals
        if ns_vals:
            context["nameservers"] = ns_vals
        if dns_vals:
            context["dns_values"] = dns_vals

    # Additional API families beyond Explore API (documented endpoints only).
    non_explore_payloads: List[Any] = []
    non_explore_available_apis: List[str] = []
    non_explore_api_results: Dict[str, Any] = {}
    non_explore_unavailable_attempts: List[str] = []
    non_explore_status: Dict[str, str] = {}

    asn_for_probe = ""
    for candidate in (
        context.get("answer_asn"),
        context.get("asn"),
        top_row.get("answer_asn") if isinstance(top_row, dict) else None,
        top_row.get("asn") if isinstance(top_row, dict) else None,
    ):
        normalized = _silentpush_normalize_asn(candidate)
        if normalized:
            asn_for_probe = normalized
            break

    if not asn_for_probe:
        asn_for_probe = _silentpush_extract_asn_for_probe([payload, *enrichment_payloads])

    normalized_url = ioc if ioc_type == "url" else f"http://{indicator}"

    family_specs: List[Dict[str, Any]] = []

    if SILENTPUSH_ENABLE_CONTEXT_GRAPH_SEARCH:
        family_specs.append({
            "api": "context_graph_search",
            "variants": _silentpush_context_graph_variants(indicator, ioc, ioc_type),
        })

    if SILENTPUSH_ENABLE_ENRICH_INDICATOR:
        enrich_indicator_type = "ipv4" if query_type == "ip" else "domain"
        enrich_variants = []
        if query_type == "ip":
            enrich_variants = [
                _silentpush_variant("GET", f"/explore/enrich/{enrich_indicator_type}/{indicator_enc}", params={"explain": 1}, base_url=SILENTPUSH_API_BASE_URL),
                _silentpush_variant("GET", f"/explore/ipv4/enrichment/{indicator_enc}", base_url=SILENTPUSH_API_BASE_URL),
            ]
        else:
            enrich_variants = [
                _silentpush_variant("GET", f"/explore/enrich/{enrich_indicator_type}/{indicator_enc}", params={"explain": 1}, base_url=SILENTPUSH_API_BASE_URL),
                _silentpush_variant("GET", f"/explore/domain/enrichment/{indicator_enc}", base_url=SILENTPUSH_API_BASE_URL),
            ]
        family_specs.append({"api": "enrich_indicator", "variants": enrich_variants})

    if SILENTPUSH_ENABLE_PADNS_LOOKUP:
        padns_variants = [
            _silentpush_variant("GET", "/explore/padns/lookup/query/a", params={"query": indicator}, base_url=SILENTPUSH_API_BASE_URL),
            _silentpush_variant("GET", f"/explore/padns/lookup/query/a/{indicator_enc}", params={"limit": 1000}, base_url=SILENTPUSH_API_BASE_URL),
        ]
        if query_type == "domain":
            for record in ("aaaa", "mx", "ns", "txt", "cname"):
                padns_variants.append(
                    _silentpush_variant(
                        "GET",
                        f"/explore/padns/lookup/query/{record}/{indicator_enc}",
                        params={"limit": 1000},
                        base_url=SILENTPUSH_API_BASE_URL,
                    )
                )
        family_specs.append({"api": "padns_lookup", "variants": padns_variants})

    if SILENTPUSH_ENABLE_BULK_ENRICH:
        if query_type == "domain":
            family_specs.append({
                "api": "bulk_summary_domain",
                "variants": [
                    _silentpush_variant("POST", "/explore/bulk/summary/domain", payload={"domains": [indicator]}, base_url=SILENTPUSH_API_BASE_URL),
                    _silentpush_variant("POST", "/explore/bulk/summary/domain", payload={"query": [indicator]}, base_url=SILENTPUSH_API_BASE_URL),
                    _silentpush_variant("POST", "/explore/bulk/summary/domain", payload={"indicators": [indicator]}, base_url=SILENTPUSH_API_BASE_URL),
                ],
            })
        elif query_type == "ip":
            bulk_ip_version = "ipv6" if ":" in indicator else "ipv4"
            family_specs.append({
                "api": "bulk_ip2asn",
                "variants": [
                    _silentpush_variant("POST", f"/explore/bulk/ip2asn/{bulk_ip_version}", payload={"ips": [indicator]}, base_url=SILENTPUSH_API_BASE_URL),
                    _silentpush_variant("POST", f"/explore/bulk/ip2asn/{bulk_ip_version}", payload={"query": [indicator]}, base_url=SILENTPUSH_API_BASE_URL),
                    _silentpush_variant("POST", f"/explore/bulk/ip2asn/{bulk_ip_version}", payload={"indicators": [indicator]}, base_url=SILENTPUSH_API_BASE_URL),
                ],
            })

    if SILENTPUSH_ENABLE_THREAT_CHECK:
        threat_check_type = "url" if ioc_type == "url" else ("ip" if query_type == "ip" else "domain")
        threat_check_query = ioc if ioc_type == "url" else indicator
        family_specs.append({
            "api": "threat_check",
            "variants": [
                _silentpush_variant(
                    "GET",
                    "/v1/",
                    params={
                        "t": threat_check_type,
                        "d": SILENTPUSH_DATA_SOURCE or "iofa",
                        "u": SILENTPUSH_ACCESS_KEY,
                        "q": threat_check_query,
                    },
                    base_url=SILENTPUSH_THREAT_CHECK_BASE_URL,
                ),
            ],
        })

    if SILENTPUSH_ENABLE_LIVE_SCAN and ioc_type == "url":
        family_specs.append({
            "api": "live_scan",
            "variants": [
                _silentpush_variant(
                    "GET",
                    "/live-scan/scan-on-demand",
                    params={"url": normalized_url},
                    base_url=SILENTPUSH_V2_API_BASE_URL,
                ),
                _silentpush_variant(
                    "POST",
                    "/live-scan/scan-on-demand",
                    payload={"url": normalized_url},
                    base_url=SILENTPUSH_V2_API_BASE_URL,
                ),
            ],
        })
    elif SILENTPUSH_ENABLE_LIVE_SCAN:
        non_explore_status["live_scan"] = "skipped_non_url_ioc"

    if SILENTPUSH_ENABLE_THREAT_RANKING:
        family_specs.append({
            "api": "threat_ranking",
            "variants": [
                _silentpush_variant(
                    "GET",
                    "/iocs/threat-ranking",
                    params={"query": indicator, "query_type": query_type},
                    base_url=SILENTPUSH_V2_API_BASE_URL,
                ),
                _silentpush_variant(
                    "POST",
                    "/iocs/threat-ranking",
                    payload={"query": indicator, "query_type": query_type},
                    base_url=SILENTPUSH_V2_API_BASE_URL,
                ),
            ],
        })

    if query_type == "ip" and asn_for_probe:
        asn_enc = quote(asn_for_probe, safe="")
        family_specs.extend([
            {
                "api": "asn_ip_reputation",
                "variants": [
                    _silentpush_variant(
                        "GET",
                        f"/explore/ipreputation/asn/{asn_enc}",
                        params={"explain": 1},
                        base_url=SILENTPUSH_API_BASE_URL,
                    ),
                ],
            },
            {
                "api": "asn_takedown_reputation",
                "variants": [
                    _silentpush_variant(
                        "GET",
                        f"/explore/takedownreputation/asn/{asn_enc}",
                        params={"explain": 1},
                        base_url=SILENTPUSH_API_BASE_URL,
                    ),
                ],
            },
            {
                "api": "asn_ip_reputation_history",
                "variants": [
                    _silentpush_variant(
                        "GET",
                        f"/explore/ipreputation/history/asn/{asn_enc}",
                        params={"limit": 3, "explain": 1},
                        base_url=SILENTPUSH_API_BASE_URL,
                    ),
                ],
            },
            {
                "api": "asn_takedown_reputation_history",
                "variants": [
                    _silentpush_variant(
                        "GET",
                        f"/explore/takedownreputation/history/asn/{asn_enc}",
                        params={"limit": 3, "explain": 1},
                        base_url=SILENTPUSH_API_BASE_URL,
                    ),
                ],
            },
        ])

    for spec in family_specs[:SILENTPUSH_MAX_NON_EXPLORE_CALLS]:
        probe, unavailable = await _silentpush_probe_api_family(session, headers, spec["api"], spec["variants"])
        if probe:
            api_name = probe.get("api")
            non_explore_available_apis.append(api_name)
            non_explore_status[api_name] = "available"
            non_explore_api_results[api_name] = {
                "endpoint": probe.get("endpoint"),
                "method": probe.get("method"),
                "response": probe.get("payload_compact"),
            }
            payload_obj = probe.get("payload")
            if isinstance(payload_obj, (dict, list)):
                non_explore_payloads.append(payload_obj)
            continue

        status_code = "unknown"
        if unavailable:
            last = unavailable[-1]
            status_code = last.rsplit(":", 1)[-1]
            if SILENTPUSH_VERBOSE_NON_EXPLORE_ERRORS:
                non_explore_unavailable_attempts.extend(unavailable)
        non_explore_status[spec["api"]] = f"unavailable_{status_code}"

    if non_explore_available_apis:
        context["non_explore_available_apis"] = non_explore_available_apis
    if non_explore_status:
        context["non_explore_status"] = non_explore_status
    if non_explore_unavailable_attempts:
        context["non_explore_unavailable_attempts"] = non_explore_unavailable_attempts
    if non_explore_api_results:
        context["non_explore_api_results"] = non_explore_api_results

    if asn_for_probe:
        context["asn_lookup"] = asn_for_probe

    if "asn_ip_reputation" in non_explore_api_results:
        context["asn_ip_reputation"] = non_explore_api_results["asn_ip_reputation"].get("response")
    if "asn_takedown_reputation" in non_explore_api_results:
        context["asn_takedown_reputation"] = non_explore_api_results["asn_takedown_reputation"].get("response")
    if "asn_ip_reputation_history" in non_explore_api_results:
        context["asn_ip_reputation_history"] = non_explore_api_results["asn_ip_reputation_history"].get("response")
    if "asn_takedown_reputation_history" in non_explore_api_results:
        context["asn_takedown_reputation_history"] = non_explore_api_results["asn_takedown_reputation_history"].get("response")
    if "threat_check" in non_explore_api_results:
        threat_check_response = non_explore_api_results["threat_check"].get("response")
        context["threat_check_response"] = threat_check_response
        threat_check_listed = _silentpush_parse_threat_check_listed(threat_check_response)
        if threat_check_listed is not None:
            context["threat_check_listed"] = threat_check_listed
            if threat_check_listed:
                is_listed = True
                sp_verdict = "Malicious"
                score = max(score, 16)
                listed_txt = f"{listed_txt}; Threat Check listed=true"

    if query_type == "ip":
        padns_payload = None
        if "padns_lookup" in non_explore_api_results:
            padns_payload = non_explore_api_results["padns_lookup"].get("response")
        signal, domain_count, domain_sample, signal_reason = _silentpush_bulletproof_hosting_signal_from_padns(padns_payload)
        context["bulletproof_hosting_likelihood"] = signal
        context["bulletproof_hosting_reason"] = signal_reason
        if domain_count is not None:
            context["padns_reverse_a_domains_count"] = domain_count
        if domain_sample:
            context["padns_reverse_a_domains_sample"] = domain_sample
        if signal in {"likely", "suspected"}:
            context["bulletproof_hosting_flag"] = True

    if non_explore_payloads:
        threat_rank_vals = _silentpush_collect_values(
            non_explore_payloads,
            ("threat_rank", "rank", "priority", "risk_rank"),
            limit=3,
        )
        live_scan_status_vals = _silentpush_collect_values(
            non_explore_payloads,
            ("scan_status", "scan_state", "status", "state", "verdict"),
            limit=4,
        )
        context_graph_count_vals = _silentpush_collect_values(
            non_explore_payloads,
            ("matches", "result_count", "total", "count", "nodes"),
            limit=4,
        )

        if threat_rank_vals:
            context["threat_rank"] = threat_rank_vals[0]
        if live_scan_status_vals:
            context["live_scan_status"] = live_scan_status_vals[0]
        if context_graph_count_vals:
            context["context_graph_hits"] = context_graph_count_vals[0]

    all_payloads: List[Any] = [payload]
    all_payloads.extend(enrichment_payloads)
    all_payloads.extend(non_explore_payloads)
    extra_value_specs = {
        "answer": ("answer", "resolved_ip"),
        "answer_asn": ("answer_asn",),
        "answer_as_name": ("answer_as_name",),
        "country_code": ("country_code",),
        "subnet": ("subnet", "cidr", "netblock", "prefix"),
        "asn_reputation": ("asn_reputation",),
        "active_ips": ("active_ips",),
        "active_subnets": ("active_subnets",),
        "ips_in_asn": ("ips_in_asn",),
        "ips_listed": ("ips_listed",),
        "registration_date": ("registration_date",),
        "last_changed_date": ("last_changed_date",),
        "expiration_date": ("expiration_date",),
        "whois_server": ("whois_server",),
        "url": ("url",),
    }
    for output_key, key_terms in extra_value_specs.items():
        if output_key in context:
            continue
        vals = _silentpush_collect_values(all_payloads, key_terms, limit=2)
        if vals:
            context[output_key] = vals[0]

    out = {
        "source": "silentpush",
        "endpoint": endpoint,
        "query": indicator,
        "query_type": query_type,
        "data_source": "explore-api+additional-families",
        "sp_risk_score": risk_score,
        "sp_verdict": sp_verdict,
        "is_listed": is_listed,
        "listed_txt": listed_txt,
        "score": score,
        **context,
    }
    if api_usage:
        out["api_usage"] = api_usage
    return out


async def query_spur(session, ioc, ioc_type):
    if ioc_type != "ip":
        return {"source": "spur", "score": 0, "error": "unsupported_ioc_type"}
    if not SPUR_API_KEY:
        return {"source": "spur", "score": 0, "error": "no_api_key"}

    url = f"https://api.spur.us/v2/context/{ioc}"
    headers = {"Token": SPUR_API_KEY}
    params = {}
    spur_dt = _spur_sanitized_dt()
    if spur_dt:
        params["dt"] = spur_dt
    if SPUR_USE_MAXMIND_GEO:
        params["mmgeo"] = "1"

    status, data, request_error, response_headers = await _request_json_with_retry(
        session,
        "GET",
        url,
        headers=headers,
        params=params or None,
        return_headers=True,
    )
    api_usage = _extract_api_usage(response_headers)
    if request_error:
        out = {"source": "spur", "score": 0, "error": request_error}
        if api_usage:
            out["api_usage"] = api_usage
        return out
    if status == 401:
        out = {"source": "spur", "score": 0, "error": "invalid_api_key"}
        if api_usage:
            out["api_usage"] = api_usage
        return out
    if status == 403:
        out = {
            "source": "spur",
            "score": 0,
            "error": "no_context_api_access",
            "hint": "Token is valid but does not include Context API access.",
        }
        if api_usage:
            out["api_usage"] = api_usage
        return out
    if status == 429:
        out = {"source": "spur", "score": 0, "error": "rate_limited"}
        if api_usage:
            out["api_usage"] = api_usage
        return out
    if status == 400:
        out = {"source": "spur", "score": 0, "error": "invalid_query"}
        if api_usage:
            out["api_usage"] = api_usage
        return out
    if status == 500:
        out = {"source": "spur", "score": 0, "error": "http_500"}
        if api_usage:
            out["api_usage"] = api_usage
        return out
    if status and status >= 400:
        out = {"source": "spur", "score": 0, "error": f"http_{status}"}
        if api_usage:
            out["api_usage"] = api_usage
        return out

    infrastructure = str(data.get("infrastructure", "") or "")
    risks = [str(r) for r in (data.get("risks") or []) if r]
    tunnels = data.get("tunnels") or []
    services = [str(s) for s in (data.get("services") or []) if s]
    client = data.get("client") or {}
    concentration = client.get("concentration") or {}
    proxies = [str(p) for p in (client.get("proxies") or []) if p]
    client_behaviors = [str(b) for b in (client.get("behaviors") or []) if b]
    client_types = [str(t) for t in (client.get("types") or []) if t]
    location = data.get("location") or {}
    ai_obj = data.get("ai") or {}

    tunnel_types = set()
    tunnel_operators = set()
    tunnel_tags = set()
    for tunnel in tunnels:
        if not isinstance(tunnel, dict):
            continue
        for key in ("type", "tunnelType", "classification"):
            val = tunnel.get(key)
            if val:
                tunnel_types.add(str(val))
        for key in ("operator", "organization", "name", "service"):
            val = tunnel.get(key)
            if val:
                tunnel_operators.add(str(val))
        for key in ("tag", "serviceTag"):
            val = tunnel.get(key)
            if val:
                tunnel_tags.add(str(val))

    score = 0
    if risks:
        score += min(12, len(risks) * 2)
    if tunnels:
        score += 5
    if proxies:
        score += 3
    score = min(20, score)
    spur_verdict = _spur_verdict_from_signals(risks, len(tunnels), len(proxies))

    tag_metadata: Dict[str, Dict[str, Any]] = {}
    if SPUR_ENABLE_TAG_METADATA and SPUR_MAX_TAG_METADATA > 0 and tunnel_tags:
        for tag in sorted(tunnel_tags)[:SPUR_MAX_TAG_METADATA]:
            md = await _spur_fetch_tag_metadata(session, tag)
            if md:
                tag_metadata[tag] = md

    result = {
        "source": "spur",
        "score": score,
        "spur_verdict": spur_verdict,
        "ip": data.get("ip", ioc),
        "infrastructure": infrastructure,
        "risks": risks[:10],
        "services": services[:10],
        "tunnel_count": len(tunnels),
        "proxy_count": len(proxies),
        "asn": (data.get("as") or {}).get("number"),
        "as_organization": (data.get("as") or {}).get("organization"),
        "organization": data.get("organization"),
        "country": location.get("country"),
        "city": location.get("city"),
        "balance_remaining": response_headers.get("x-balance-remaining"),
        "result_dt": response_headers.get("x-result-dt"),
        "concentration_country": concentration.get("country"),
        "concentration_state": concentration.get("state"),
        "concentration_city": concentration.get("city"),
        "concentration_density": concentration.get("density"),
        "concentration_skew": concentration.get("skew"),
        "concentration_geohash": concentration.get("geohash"),
        "client_count": client.get("count"),
        "client_countries": client.get("countries"),
        "client_spread": client.get("spread"),
        "client_types": client_types[:10],
        "tunnel_types": sorted(tunnel_types)[:10],
        "tunnel_operators": sorted(tunnel_operators)[:10],
        "tunnel_tags": sorted(tunnel_tags)[:10],
        "mmgeo_enabled": SPUR_USE_MAXMIND_GEO,
    }
    if api_usage:
        result["api_usage"] = api_usage
    if spur_dt:
        result["query_dt"] = spur_dt
    if client_behaviors:
        result["client_behaviors"] = client_behaviors[:10]
    if ai_obj:
        result["ai"] = ai_obj
    if tag_metadata:
        result["tag_metadata"] = tag_metadata
    return result


def get_spur_token_status(timeout: int = 4) -> Dict[str, Any]:
    if not SPUR_API_KEY:
        return {"active": False, "error": "no_api_key"}

    req = urllib.request.Request(
        "https://api.spur.us/status",
        headers={"Token": SPUR_API_KEY},
        method="GET",
    )
    try:
        with urllib.request.urlopen(req, timeout=timeout) as resp:
            raw = resp.read().decode("utf-8", errors="replace")
            data = json.loads(raw) if raw else {}
            return {
                "active": bool(data.get("active")),
                "queries_remaining": data.get("queriesRemaining"),
                "service_tier": data.get("serviceTier"),
            }
    except urllib.error.HTTPError as e:
        if e.code == 401:
            return {"active": False, "error": "invalid_api_key"}
        if e.code == 403:
            return {"active": False, "error": "forbidden"}
        if e.code == 429:
            return {"active": False, "error": "rate_limited"}
        return {"active": False, "error": f"http_{e.code}"}
    except Exception as e:
        return {"active": False, "error": str(e)}


def get_reversinglabs_token_status(timeout: int = 4) -> Dict[str, Any]:
    if not RL_SPECTRA_BASE_URL:
        return {"active": False, "error": "no_base_url"}
    if not RL_SPECTRA_TOKEN:
        return {"active": False, "error": "no_api_key"}

    req = urllib.request.Request(
        f"{RL_SPECTRA_BASE_URL}/api/license/v1/",
        headers=_rl_auth_headers(),
        method="GET",
    )
    try:
        with urllib.request.urlopen(req, timeout=timeout) as resp:
            raw = resp.read().decode("utf-8", errors="replace")
            data = json.loads(raw) if raw else {}
            return {
                "active": True,
                "status": data.get("status") or data.get("service_status") or "ok",
                "license": data.get("license") if isinstance(data, dict) else None,
            }
    except urllib.error.HTTPError as e:
        if e.code == 401:
            return {"active": False, "error": "invalid_api_key"}
        if e.code == 403:
            return {"active": False, "error": "forbidden"}
        if e.code == 429:
            return {"active": False, "error": "rate_limited"}
        if e.code == 404:
            return {"active": False, "error": "not_found"}
        return {"active": False, "error": f"http_{e.code}"}
    except Exception as e:
        return {"active": False, "error": str(e)}


def _rl_auth_headers() -> Dict[str, str]:
    token = (RL_SPECTRA_TOKEN or "").strip()
    return {
        "Authorization": f"Token {token}",
        "Token": token,
        "Accept": "application/json",
    }


def _rl_leaf_items(value, prefix: str = "", depth: int = 0):
    if depth > 4:
        return
    if isinstance(value, dict):
        for k, v in value.items():
            key = f"{prefix}.{k}" if prefix else str(k)
            yield from _rl_leaf_items(v, key, depth + 1)
        return
    if isinstance(value, list):
        for idx, item in enumerate(value[:10]):
            key = f"{prefix}[{idx}]"
            yield from _rl_leaf_items(item, key, depth + 1)
        return
    yield prefix, value


def _rl_extract_signal(data: Dict[str, Any]) -> Dict[str, Any]:
    malicious_terms = ("malicious", "phishing", "malware", "c2", "botnet", "suspicious")
    benign_terms = ("benign", "clean", "safe", "trusted")

    leaf_items = list(_rl_leaf_items(data))
    classification = ""
    threat_level = ""
    risk_score = None
    malicious_hits = 0
    benign_hits = 0

    for key, val in leaf_items:
        key_l = key.lower()
        if isinstance(val, (int, float)):
            if risk_score is None and any(k in key_l for k in ("risk_score", "threat_score", "score")):
                risk_score = float(val)
            if "malicious" in key_l and float(val) > 0:
                malicious_hits += 1
            continue

        text = str(val).strip()
        if not text:
            continue
        text_l = text.lower()

        if not classification and "class" in key_l:
            classification = text
        if not threat_level and "threat" in key_l and "level" in key_l:
            threat_level = text

        if any(term in text_l for term in malicious_terms):
            malicious_hits += 1
        if any(term in text_l for term in benign_terms):
            benign_hits += 1

    score = 0
    cls_l = classification.lower()
    if any(term in cls_l for term in ("malicious", "known_malicious", "high")):
        score += 14
    elif any(term in cls_l for term in ("suspicious", "medium", "unknown")):
        score += 8
    elif any(term in cls_l for term in benign_terms):
        score += 0

    if risk_score is not None:
        if risk_score <= 1:
            score += round(risk_score * 10)
        else:
            score += round(min(risk_score, 100) / 10)

    score += min(8, malicious_hits * 2)
    if benign_hits > 0 and malicious_hits == 0:
        score = max(0, score - 4)

    tp_stats = ((data.get("third_party_reputations") or {}).get("statistics") or {}) if isinstance(data, dict) else {}
    tp_malicious = int(tp_stats.get("malicious") or 0)
    tp_suspicious = int(tp_stats.get("suspicious") or 0)
    tp_clean = int(tp_stats.get("clean") or 0)
    tp_total = int(tp_stats.get("total") or 0)

    threat_l = threat_level.lower()
    if ("malicious" in cls_l) or ("high" in threat_l) or tp_malicious > 0 or malicious_hits > 0:
        rl_verdict = "Malicious"
    elif ("suspicious" in cls_l) or ("medium" in threat_l) or tp_suspicious > 0:
        rl_verdict = "Suspicious"
    elif ("benign" in cls_l) or (tp_total > 0 and tp_malicious == 0 and tp_suspicious == 0 and tp_clean > 0):
        rl_verdict = "Benign"
    else:
        rl_verdict = "Unknown"

    detection_ratio = None
    if tp_total > 0:
        detection_ratio = round((tp_malicious + tp_suspicious) / tp_total, 3)

    return {
        "classification": classification,
        "threat_level": threat_level,
        "risk_score": risk_score,
        "malicious_signals": malicious_hits,
        "rl_verdict": rl_verdict,
        "third_party_malicious": tp_malicious,
        "third_party_suspicious": tp_suspicious,
        "third_party_clean": tp_clean,
        "third_party_total": tp_total,
        "third_party_detection_ratio": detection_ratio,
        "score": min(20, max(0, score)),
    }


def _rl_extract_collection_values(payload: Any) -> List[Any]:
    if isinstance(payload, list):
        return payload
    if isinstance(payload, dict):
        for key in ("results", "items", "data", "objects", "records", "list"):
            val = payload.get(key)
            if isinstance(val, list):
                return val
    return []


def _rl_extract_highlights(data: Dict[str, Any]) -> Dict[str, Any]:
    keywords = (
        "classification", "threat", "risk", "reputation", "confidence", "severity", "status",
        "family", "campaign", "actor", "malware", "phishing", "first_seen", "last_seen",
    )
    highlights: Dict[str, Any] = {}
    tags: List[str] = []

    for key, val in _rl_leaf_items(data):
        if len(highlights) >= 18:
            break
        key_l = key.lower()
        if not any(k in key_l for k in keywords):
            continue
        if isinstance(val, (str, int, float, bool)):
            text = str(val).strip()
            if text:
                highlights[key] = val
        if "tag" in key_l and isinstance(val, str) and val.strip():
            tags.append(val.strip())

    if tags:
        highlights["top_tags"] = sorted(set(tags))[:10]
    return highlights


def _rl_extract_whois_context(data: Dict[str, Any]) -> Dict[str, Any]:
    leaf_items = list(_rl_leaf_items(data))

    single_value_aliases = {
        "registrar": ("registrar",),
        "registrant": ("registrant", "registrant_name", "registrant_org", "registrant_organization"),
        "organization": ("organization", "org", "owner", "registrant_org", "registrant_organization"),
        "country": ("country", "country_code"),
        "created": ("creation_date", "created", "created_at", "registered_on", "registration_date"),
        "updated": ("updated", "updated_at", "updated_date", "last_updated"),
        "expires": ("expiration_date", "expires", "expires_at", "expiry_date", "expiration"),
        "asn": ("asn", "as_number"),
        "as_owner": ("as_owner", "as_name", "as_org", "as_organization", "asn_org"),
        "network": ("network", "cidr", "netblock", "network_range"),
        "registry": ("registry", "rir"),
    }
    list_value_aliases = {
        "nameservers": ("nameserver", "nameservers", "name_server"),
    }

    context: Dict[str, Any] = {}
    list_values: Dict[str, List[str]] = {key: [] for key in list_value_aliases}

    for key, val in leaf_items:
        if isinstance(val, (dict, list, tuple, set)):
            continue
        text = str(val).strip()
        if not text:
            continue

        key_l = key.lower()
        normalized_key = key_l.replace("[", ".").replace("]", "")

        for output_key, aliases in single_value_aliases.items():
            if output_key in context:
                continue
            if any(
                normalized_key.endswith(alias)
                or f".{alias}." in f".{normalized_key}."
                for alias in aliases
            ):
                context[output_key] = text
                break

        for output_key, aliases in list_value_aliases.items():
            if any(
                normalized_key.endswith(alias)
                or f".{alias}." in f".{normalized_key}."
                for alias in aliases
            ):
                if text not in list_values[output_key]:
                    list_values[output_key].append(text)

    for output_key, values in list_values.items():
        if values:
            context[output_key] = values[:10]

    return context


def _rl_summarize_collection(payload: Any) -> Dict[str, Any]:
    items = _rl_extract_collection_values(payload)
    if not items:
        return {"count": 0, "sample": []}

    sample: List[str] = []
    for item in items[:5]:
        if isinstance(item, str):
            sample.append(item)
            continue
        if isinstance(item, dict):
            for candidate_key in ("value", "url", "domain", "ip", "filename", "name"):
                value = item.get(candidate_key)
                if value:
                    sample.append(str(value))
                    break
    return {"count": len(items), "sample": sample[:5]}


async def query_reversinglabs(session, ioc, ioc_type):
    if not RL_SPECTRA_BASE_URL:
        return {"source": "reversinglabs", "score": 0, "error": "no_base_url"}
    if not RL_SPECTRA_TOKEN:
        return {"source": "reversinglabs", "score": 0, "error": "no_api_key"}

    headers = _rl_auth_headers()

    endpoint = ""
    status = None
    data = {}
    request_error = None
    response_headers: Dict[str, Any] = {}

    # Try endpoint variants because Spectra deployments can differ on exact URL shape.
    attempts = []
    if ioc_type == "ip":
        attempts = [
            ("GET", f"/api/network-threat-intel/ip/{ioc}/report", None),
            ("GET", f"/api/network-threat-intel/ip/{ioc}/report/", None),
            ("GET", f"/api/network-threat-intel/ip/{ioc}/urls", None),
            ("GET", f"/api/network-threat-intel/ip/{ioc}/urls/", None),
            ("GET", f"/api/network-threat-intel/ip/{ioc}/resolutions", None),
            ("GET", f"/api/network-threat-intel/ip/{ioc}/resolutions/", None),
        ]
    elif ioc_type == "domain":
        attempts = [
            ("GET", f"/api/network-threat-intel/domain/{ioc}", None),
            ("GET", f"/api/network-threat-intel/domain/{ioc}/", None),
        ]
    elif ioc_type == "url":
        attempts = [
            ("GET", "/api/network-threat-intel/url", {"url": ioc}),
            ("GET", "/api/network-threat-intel/url/", {"url": ioc}),
            ("GET", "/api/network-threat-intel/url", {"value": ioc}),
            ("GET", "/api/network-threat-intel/url/", {"value": ioc}),
            ("GET", "/api/network-threat-intel/url", {"q": ioc}),
            ("GET", "/api/network-threat-intel/url/", {"q": ioc}),
        ]
    elif ioc_type == "hash":
        attempts = [
            ("GET", f"/api/samples/v3/{ioc}/classification", None),
            ("GET", f"/api/samples/v3/{ioc}/classification/", None),
        ]
    else:
        return {"source": "reversinglabs", "score": 0, "error": "unsupported_ioc_type"}

    last_endpoint = ""
    all_404 = True
    for method, ep, params in attempts:
        last_endpoint = ep
        status, data, request_error, headers_out = await _request_json_with_retry(
            session,
            method,
            f"{RL_SPECTRA_BASE_URL}{ep}",
            headers=headers,
            params=params,
            return_headers=True,
        )
        response_headers = headers_out or response_headers
        endpoint = ep

        if request_error:
            break
        if status == 404:
            continue
        all_404 = False
        # Keep trying URL shape variations only for bad query parameter mappings.
        if ioc_type == "url" and status == 400:
            continue
        break

    if all_404:
        out = {
            "source": "reversinglabs",
            "score": 0,
            "error": "not_found",
            "endpoint": last_endpoint,
            "hint": "No supported Network Threat Intel endpoint variant was found on this Spectra instance.",
        }
        api_usage = _extract_api_usage(response_headers)
        if api_usage:
            out["api_usage"] = api_usage
        return out

    api_usage = _extract_api_usage(response_headers)
    if request_error:
        out = {"source": "reversinglabs", "score": 0, "error": request_error, "endpoint": endpoint}
        if api_usage:
            out["api_usage"] = api_usage
        return out
    if status == 401:
        out = {"source": "reversinglabs", "score": 0, "error": "invalid_api_key", "endpoint": endpoint}
        if api_usage:
            out["api_usage"] = api_usage
        return out
    if status == 403:
        out = {"source": "reversinglabs", "score": 0, "error": "forbidden", "endpoint": endpoint}
        if api_usage:
            out["api_usage"] = api_usage
        return out
    if status == 404:
        out = {"source": "reversinglabs", "score": 0, "error": "not_found", "endpoint": endpoint}
        if api_usage:
            out["api_usage"] = api_usage
        return out
    if status == 429:
        out = {"source": "reversinglabs", "score": 0, "error": "rate_limited", "endpoint": endpoint}
        if api_usage:
            out["api_usage"] = api_usage
        return out
    if status and status >= 400:
        out = {"source": "reversinglabs", "score": 0, "error": f"http_{status}", "endpoint": endpoint}
        if api_usage:
            out["api_usage"] = api_usage
        return out

    normalized = data if isinstance(data, dict) else {"items": data, "count": len(data) if isinstance(data, list) else 0}
    signal = _rl_extract_signal(normalized)
    highlights = _rl_extract_highlights(normalized)
    whois_context = _rl_extract_whois_context(normalized)

    # Best-effort additional context for IP IOCs.
    related_urls = {"count": 0, "sample": []}
    related_resolutions = {"count": 0, "sample": []}
    related_downloads = {"count": 0, "sample": []}
    if ioc_type == "ip":
        for path, sink in (
            (f"/api/network-threat-intel/ip/{ioc}/urls/", "urls"),
            (f"/api/network-threat-intel/ip/{ioc}/resolutions/", "resolutions"),
            (f"/api/network-threat-intel/ip/{ioc}/downloaded_files/", "downloads"),
        ):
            s2, d2, e2 = await _request_json_with_retry(
                session,
                "GET",
                f"{RL_SPECTRA_BASE_URL}{path}",
                headers=headers,
            )
            if e2 or s2 in {401, 403, 404, 429}:
                continue
            summary = _rl_summarize_collection(d2)
            if sink == "urls":
                related_urls = summary
            elif sink == "resolutions":
                related_resolutions = summary
            else:
                related_downloads = summary

    out = {
        "source": "reversinglabs",
        "endpoint": endpoint,
        "classification": signal.get("classification"),
        "threat_level": signal.get("threat_level"),
        "risk_score": signal.get("risk_score"),
        "malicious_signals": signal.get("malicious_signals", 0),
        "rl_verdict": signal.get("rl_verdict"),
        "third_party_malicious": signal.get("third_party_malicious"),
        "third_party_suspicious": signal.get("third_party_suspicious"),
        "third_party_clean": signal.get("third_party_clean"),
        "third_party_total": signal.get("third_party_total"),
        "third_party_detection_ratio": signal.get("third_party_detection_ratio"),
        "highlights": highlights,
        "registrar": whois_context.get("registrar"),
        "registrant": whois_context.get("registrant"),
        "organization": whois_context.get("organization"),
        "country": whois_context.get("country"),
        "created": whois_context.get("created"),
        "updated": whois_context.get("updated"),
        "expires": whois_context.get("expires"),
        "asn": whois_context.get("asn"),
        "as_owner": whois_context.get("as_owner"),
        "network": whois_context.get("network"),
        "registry": whois_context.get("registry"),
        "nameservers": whois_context.get("nameservers", []),
        "related_urls_count": related_urls.get("count", 0),
        "related_urls_sample": related_urls.get("sample", []),
        "resolutions_count": related_resolutions.get("count", 0),
        "resolutions_sample": related_resolutions.get("sample", []),
        "downloaded_files_count": related_downloads.get("count", 0),
        "downloaded_files_sample": related_downloads.get("sample", []),
        "score": signal.get("score", 0),
    }
    if api_usage:
        out["api_usage"] = api_usage
    return out


# ----------------------------
# SCORING (proportional)
# ----------------------------
def calculate_score(results):
    total = 0.0
    reasons = []

    for r in results:
        if r.get("error"):
            continue

        if r["source"] == "alienvault":
            pulses = r.get("pulses", 0)
            if pulses > 0:
                # 1 pulse ≈ 2.5 pts; caps at 25 pts at 10+ pulses
                pts = round(min(pulses / 10 * 25, 25))
                total += pts
                label = "pulse" if pulses == 1 else "pulses"
                reasons.append(f"AlienVault: {pulses} threat {label} (+{pts}pts)")

        if r["source"] == "abuseipdb":
            abuse_score = r.get("abuse_score", 0)
            if abuse_score > 0:
                # Raw 0–100 confidence mapped to 0–20 pts
                pts = round(abuse_score / 100 * 20)
                total += pts
                reasons.append(f"AbuseIPDB: confidence {abuse_score}/100 (+{pts}pts)")

        if r["source"] == "threatfox":
            pts = r.get("score", 0)
            if pts > 0:
                total += pts
                matches = r.get("matches", 0)
                max_confidence = r.get("max_confidence", 0)
                reasons.append(f"ThreatFox: {matches} match(es), confidence {max_confidence}/100 (+{pts}pts)")

        if r["source"] == "silentpush":
            risk_score = r.get("sp_risk_score", 0)
            pts = r.get("score", 0)
            if pts > 0:
                total += pts
                extra = ""
                available = r.get("available_endpoints") or []
                api_families = r.get("non_explore_available_apis") or []
                if available:
                    extra = f", enrichment endpoints={len(available)}"
                if api_families:
                    extra = f"{extra}, additional api families={len(api_families)}"
                reasons.append(
                    f"Silent Push Explore: risk score {risk_score}/100 via {r.get('query_type', 'indicator')} lookup{extra} (+{pts}pts)"
                )
            bph_signal = str(r.get("bulletproof_hosting_likelihood") or "").lower()
            if bph_signal in {"likely", "suspected"}:
                reverse_domains = r.get("padns_reverse_a_domains_count")
                reasons.append(
                    f"Silent Push PADNS: reverse A domain volume={reverse_domains} indicates {bph_signal} bulletproof-hosting behavior"
                )

        if r["source"] == "spur":
            pts = r.get("score", 0)
            if pts > 0:
                total += pts
                infra = r.get("infrastructure") or "unknown"
                risks = r.get("risks") or []
                reasons.append(f"Spur: infrastructure={infra}, risks={len(risks)} (+{pts}pts)")

        if r["source"] == "reversinglabs":
            pts = r.get("score", 0)
            if pts > 0:
                total += pts
                verdict = r.get("rl_verdict") or "Unknown"
                detected = r.get("third_party_malicious") or 0
                suspicious = r.get("third_party_suspicious") or 0
                reasons.append(f"ReversingLabs: verdict={verdict}, third-party flags m={detected} s={suspicious} (+{pts}pts)")

    return round(total), reasons


def classify(score):
    if score >= 70:
        return "Malicious"
    elif score >= 40:
        return "Suspicious"
    elif score >= 15:
        return "Low Risk"
    return "Benign"


# ----------------------------
# EARLY EXIT
# ----------------------------
def early_exit(results):
    for r in results:
        if r.get("error"):
            continue
        if r["source"] == "silentpush" and r.get("sp_verdict") == "Malicious":
            return True, f"Silent Push returned high-severity malicious indicators (risk score {r.get('sp_risk_score')}/100)"
        if r["source"] == "spur" and r.get("spur_verdict") == "Malicious":
            risk_count = len(r.get("risks") or [])
            return True, f"Spur returned malicious-class risk indicators ({risk_count} risk signal(s))"
        if r["source"] == "reversinglabs" and r.get("rl_verdict") == "Malicious":
            return True, "ReversingLabs returned a malicious verdict"
        if r["source"] == "abuseipdb" and r.get("abuse_score", 0) > 90:
            return True, f"Very high AbuseIPDB confidence score ({r['abuse_score']}/100)"
        if r["source"] == "threatfox" and r.get("matches", 0) > 0 and r.get("max_confidence", 0) >= 90:
            return True, f"High ThreatFox confidence ({r['max_confidence']}/100 across {r['matches']} matches)"
    return False, ""


# ----------------------------
# MAIN ANALYSIS
# ----------------------------
async def analyze_with_session(
    session: aiohttp.ClientSession,
    ioc: str,
    enabled_sources: List[str] = None,
):
    start = time.time()
    ioc = ioc.strip()

    if ioc in ALLOWLIST:
        return {"ioc": ioc, "verdict": "Benign", "score": 0, "reason": "Allowlisted", "time_taken": 0}

    ioc_type = detect_type(ioc)

    enabled = set(enabled_sources or [
        "virustotal",
        "alienvault",
        "threatfox",
        "silentpush",
        "reversinglabs",
        "abuseipdb",
        "spur",
    ])

    tasks = []
    if "virustotal" in enabled:
        tasks.append(query_virustotal(session, ioc, ioc_type))
    if "alienvault" in enabled:
        tasks.append(query_alienvault(session, ioc, ioc_type))
    if "threatfox" in enabled:
        tasks.append(query_threatfox(session, ioc, ioc_type))
    if "silentpush" in enabled:
        tasks.append(query_silentpush(session, ioc, ioc_type))
    if "reversinglabs" in enabled:
        tasks.append(query_reversinglabs(session, ioc, ioc_type))
    if ioc_type == "ip":
        if "abuseipdb" in enabled:
            tasks.append(query_abuseipdb(session, ioc))
        if "spur" in enabled:
            tasks.append(query_spur(session, ioc, ioc_type))

    results = await asyncio.gather(*tasks)

    exit_flag, reason = early_exit(results)
    if exit_flag:
        return {
            "ioc": ioc, "type": ioc_type, "verdict": "Malicious",
            "score": 100, "reason": reason,
            "sources": results, "time_taken": round(time.time() - start, 2),
        }

    score, reasons = calculate_score(results)
    verdict = classify(score)

    return {
        "ioc": ioc, "type": ioc_type, "verdict": verdict,
        "score": score, "reasons": reasons,
        "sources": results, "time_taken": round(time.time() - start, 2),
    }


async def analyze(ioc: str, enabled_sources: List[str] = None):
    async with aiohttp.ClientSession() as session:
        return await analyze_with_session(session, ioc, enabled_sources=enabled_sources)


async def analyze_bulk(
    iocs: List[str],
    workers: int,
    delay: float,
    jitter: float,
    enabled_sources: List[str] = None,
):
    total = len(iocs)
    if total == 0:
        return []

    workers = max(1, workers)
    delay = max(0.0, delay)
    jitter = max(0.0, jitter)

    semaphore = asyncio.Semaphore(workers)
    pacing_lock = asyncio.Lock()
    next_allowed_start = 0.0
    output: List[Dict[str, Any]] = [None] * total  # type: ignore[list-item]

    async with aiohttp.ClientSession() as session:
        async def run_one(idx: int, ioc: str):
            nonlocal next_allowed_start
            async with semaphore:
                # Global pacing across workers to reduce bursty API traffic.
                async with pacing_lock:
                    now = time.monotonic()
                    wait_for = max(0.0, next_allowed_start - now)
                    if wait_for > 0:
                        await asyncio.sleep(wait_for)
                    spacing = delay + (random.uniform(0, jitter) if jitter > 0 else 0.0)
                    next_allowed_start = time.monotonic() + spacing

                print(f"[{idx+1}/{total}] Analyzing {ioc} ...", file=sys.stderr)
                output[idx] = await analyze_with_session(
                    session,
                    ioc,
                    enabled_sources=enabled_sources,
                )

        tasks = [asyncio.create_task(run_one(i, ioc)) for i, ioc in enumerate(iocs)]
        await asyncio.gather(*tasks)

    return output


# ----------------------------
# EXCEL INPUT READER
# ----------------------------
def _read_iocs_excel(path: str) -> List[str]:
    try:
        import openpyxl
    except ImportError:
        print("[ERROR] openpyxl is required to read Excel files. Install it with: pip install openpyxl", file=sys.stderr)
        sys.exit(1)

    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    except Exception as e:
        print(f"[ERROR] Cannot open Excel file: {e}", file=sys.stderr)
        sys.exit(1)

    ws = wb.active
    iocs = []
    for row in ws.iter_rows(min_row=1, values_only=True):
        # Read first non-empty cell in each row
        for cell in row:
            val = str(cell).strip() if cell is not None else ""
            if val and not val.startswith("#") and val.lower() not in ("ioc", "indicator", "indicators", "none"):
                iocs.append(val)
            break  # only first column
    wb.close()
    return iocs


def _prepare_iocs(iocs: List[str], max_iocs: int = 0) -> List[str]:
    """Normalize and dedupe IOCs while preserving original order."""
    deduped = []
    seen = set()
    for raw in iocs:
        ioc = raw.strip()
        if not ioc or ioc.startswith("#"):
            continue
        key = ioc.lower()
        if key in seen:
            continue
        seen.add(key)
        deduped.append(ioc)

    if max_iocs > 0:
        return deduped[:max_iocs]
    return deduped


# ----------------------------
# OUTPUT FORMATTING
# ----------------------------
EXPORT_FIELDNAMES = [
    "ioc", "type", "verdict", "score", "time_taken", "reason", "reasons",
    "source_count", "source_errors",
    "vt_malicious", "vt_suspicious", "vt_detected_by", "vt_country", "vt_continent",
    "vt_asn", "vt_as_owner", "vt_network", "vt_tags", "vt_last_analysed",
    "vt_community_reputation", "vt_community_votes", "vt_error",
    "otx_pulses", "otx_country_name", "otx_city", "otx_asn", "otx_pulse_names",
    "otx_malware_families", "otx_tags", "otx_targeted_countries", "otx_attack_ids",
    "otx_adversaries", "otx_industries_targeted", "otx_error",
    "abuse_score", "abuse_total_reports", "abuse_distinct_reporters", "abuse_country",
    "abuse_isp", "abuse_domain", "abuse_usage_type", "abuse_is_tor", "abuse_is_whitelisted",
    "abuse_last_reported_at", "abuse_hostnames", "abuse_error",
    "threatfox_matches", "threatfox_ioc_types", "threatfox_max_confidence",
    "threatfox_malware_families", "threatfox_threat_types", "threatfox_tags",
    "threatfox_first_seen", "threatfox_last_seen", "threatfox_references",
    "threatfox_malware_aliases", "threatfox_error",
    "silentpush_query", "silentpush_query_type", "silentpush_data_source", "silentpush_endpoint", "silentpush_sp_risk_score",
    "silentpush_sp_verdict", "silentpush_sp_risk_score_decider", "silentpush_answer", "silentpush_answer_asn",
    "silentpush_answer_as_name", "silentpush_country_code", "silentpush_subnet", "silentpush_asn_reputation",
    "silentpush_active_ips", "silentpush_active_subnets", "silentpush_ips_in_asn", "silentpush_ips_listed",
    "silentpush_registration_date", "silentpush_last_changed_date", "silentpush_expiration_date",
    "silentpush_whois_server", "silentpush_url", "silentpush_riskscore_row", "silentpush_riskscore_payload",
    "silentpush_asn_lookup", "silentpush_asn_ip_reputation", "silentpush_asn_takedown_reputation",
    "silentpush_asn_ip_reputation_history", "silentpush_asn_takedown_reputation_history",
    "silentpush_is_listed", "silentpush_listed_txt", "silentpush_asn", "silentpush_as_org", "silentpush_country",
    "silentpush_registrar", "silentpush_first_seen", "silentpush_last_seen", "silentpush_tags", "silentpush_nameservers",
    "silentpush_dns_values", "silentpush_available_endpoints", "silentpush_unavailable_endpoints", "silentpush_enrichment",
    "silentpush_non_explore_available_apis", "silentpush_non_explore_unavailable_attempts", "silentpush_non_explore_api_results",
    "silentpush_threat_rank", "silentpush_live_scan_status", "silentpush_context_graph_hits",
    "silentpush_bulletproof_hosting_likelihood", "silentpush_bulletproof_hosting_reason", "silentpush_bulletproof_hosting_flag",
    "silentpush_padns_reverse_a_domains_count", "silentpush_padns_reverse_a_domains_sample",
    "silentpush_error",
    "spur_ip", "spur_infrastructure", "spur_risks", "spur_services",
    "spur_tunnel_count", "spur_proxy_count", "spur_asn", "spur_as_organization",
    "spur_organization", "spur_country", "spur_city", "spur_client_behaviors",
    "spur_balance_remaining", "spur_result_dt",
    "spur_concentration_country", "spur_concentration_state", "spur_concentration_city",
    "spur_concentration_density", "spur_concentration_skew", "spur_concentration_geohash",
    "spur_client_count", "spur_client_countries", "spur_client_spread", "spur_client_types",
    "spur_tunnel_types", "spur_tunnel_operators", "spur_tunnel_tags", "spur_error",
    "rl_endpoint", "rl_verdict", "rl_classification", "rl_threat_level", "rl_risk_score", "rl_malicious_signals",
    "rl_registrar", "rl_registrant", "rl_organization", "rl_country", "rl_created", "rl_updated",
    "rl_expires", "rl_asn", "rl_as_owner", "rl_network", "rl_registry", "rl_nameservers",
    "rl_third_party_malicious", "rl_third_party_suspicious", "rl_third_party_clean", "rl_third_party_total", "rl_third_party_detection_ratio",
    "rl_highlights", "rl_related_urls_count", "rl_related_urls_sample",
    "rl_resolutions_count", "rl_resolutions_sample",
    "rl_downloaded_files_count", "rl_downloaded_files_sample", "rl_error",
    "sources_json",
]


def _join_export_value(value):
    if isinstance(value, list):
        return "; ".join(str(v) for v in value)
    if isinstance(value, dict):
        return json.dumps(value, separators=(",", ":"))
    return value


def _src_by_name(sources, name):
    return next((s for s in sources if s.get("source") == name), {})


def _flatten_export_rows(rows):
    flat_rows = []
    for row in rows:
        row_copy = {
            "ioc": row.get("ioc"),
            "type": row.get("type"),
            "verdict": row.get("verdict"),
            "score": row.get("score"),
            "time_taken": row.get("time_taken"),
            "reason": row.get("reason", ""),
            "reasons": "; ".join(row.get("reasons", [])),
        }

        sources = row.get("sources", []) or []
        vt = _src_by_name(sources, "virustotal")
        otx = _src_by_name(sources, "alienvault")
        abuse = _src_by_name(sources, "abuseipdb")
        tf = _src_by_name(sources, "threatfox")
        sp = _src_by_name(sources, "silentpush")
        spur = _src_by_name(sources, "spur")
        rl = _src_by_name(sources, "reversinglabs")

        row_copy["source_count"] = len(sources)
        row_copy["source_errors"] = "; ".join(
            f"{s.get('source')}:{s.get('error')}" for s in sources if s.get("error")
        )

        row_copy.update({
            "vt_malicious": vt.get("malicious"),
            "vt_suspicious": vt.get("suspicious"),
            "vt_detected_by": _join_export_value(vt.get("detected_by", [])),
            "vt_country": vt.get("country"),
            "vt_continent": vt.get("continent"),
            "vt_asn": vt.get("asn"),
            "vt_as_owner": vt.get("as_owner"),
            "vt_network": vt.get("network"),
            "vt_tags": _join_export_value(vt.get("tags", [])),
            "vt_last_analysed": vt.get("last_analysed"),
            "vt_community_reputation": vt.get("community_reputation"),
            "vt_community_votes": _join_export_value(vt.get("community_votes", {})),
            "vt_error": vt.get("error"),
        })

        row_copy.update({
            "otx_pulses": otx.get("pulses"),
            "otx_country_name": otx.get("country_name"),
            "otx_city": otx.get("city"),
            "otx_asn": otx.get("asn"),
            "otx_pulse_names": _join_export_value(otx.get("pulse_names", [])),
            "otx_malware_families": _join_export_value(otx.get("malware_families", [])),
            "otx_tags": _join_export_value(otx.get("tags", [])),
            "otx_targeted_countries": _join_export_value(otx.get("targeted_countries", [])),
            "otx_attack_ids": _join_export_value(otx.get("attack_ids", [])),
            "otx_adversaries": _join_export_value(otx.get("adversaries", [])),
            "otx_industries_targeted": _join_export_value(otx.get("industries_targeted", [])),
            "otx_error": otx.get("error"),
        })

        row_copy.update({
            "abuse_score": abuse.get("abuse_score"),
            "abuse_total_reports": abuse.get("total_reports"),
            "abuse_distinct_reporters": abuse.get("distinct_reporters"),
            "abuse_country": abuse.get("country"),
            "abuse_isp": abuse.get("isp"),
            "abuse_domain": abuse.get("domain"),
            "abuse_usage_type": abuse.get("usage_type"),
            "abuse_is_tor": abuse.get("is_tor"),
            "abuse_is_whitelisted": abuse.get("is_whitelisted"),
            "abuse_last_reported_at": abuse.get("last_reported_at"),
            "abuse_hostnames": _join_export_value(abuse.get("hostnames", [])),
            "abuse_error": abuse.get("error"),
        })

        row_copy.update({
            "threatfox_matches": tf.get("matches"),
            "threatfox_ioc_types": _join_export_value(tf.get("ioc_types", [])),
            "threatfox_max_confidence": tf.get("max_confidence"),
            "threatfox_malware_families": _join_export_value(tf.get("malware_families", [])),
            "threatfox_threat_types": _join_export_value(tf.get("threat_types", [])),
            "threatfox_tags": _join_export_value(tf.get("tags", [])),
            "threatfox_first_seen": tf.get("first_seen"),
            "threatfox_last_seen": tf.get("last_seen"),
            "threatfox_references": _join_export_value(tf.get("references", [])),
            "threatfox_malware_aliases": _join_export_value(tf.get("malware_aliases", [])),
            "threatfox_error": tf.get("error"),
        })

        row_copy.update({
            "silentpush_query": sp.get("query"),
            "silentpush_query_type": sp.get("query_type"),
            "silentpush_data_source": sp.get("data_source"),
            "silentpush_endpoint": sp.get("endpoint"),
            "silentpush_sp_risk_score": sp.get("sp_risk_score"),
            "silentpush_sp_verdict": sp.get("sp_verdict"),
            "silentpush_sp_risk_score_decider": sp.get("sp_risk_score_decider"),
            "silentpush_answer": sp.get("answer"),
            "silentpush_answer_asn": sp.get("answer_asn"),
            "silentpush_answer_as_name": sp.get("answer_as_name"),
            "silentpush_country_code": sp.get("country_code"),
            "silentpush_subnet": sp.get("subnet"),
            "silentpush_asn_reputation": sp.get("asn_reputation"),
            "silentpush_active_ips": sp.get("active_ips"),
            "silentpush_active_subnets": sp.get("active_subnets"),
            "silentpush_ips_in_asn": sp.get("ips_in_asn"),
            "silentpush_ips_listed": sp.get("ips_listed"),
            "silentpush_registration_date": sp.get("registration_date"),
            "silentpush_last_changed_date": sp.get("last_changed_date"),
            "silentpush_expiration_date": sp.get("expiration_date"),
            "silentpush_whois_server": sp.get("whois_server"),
            "silentpush_url": sp.get("url"),
            "silentpush_riskscore_row": _join_export_value(sp.get("riskscore_row", {})),
            "silentpush_riskscore_payload": _join_export_value(sp.get("riskscore_payload", {})),
            "silentpush_asn_lookup": sp.get("asn_lookup"),
            "silentpush_asn_ip_reputation": _join_export_value(sp.get("asn_ip_reputation", {})),
            "silentpush_asn_takedown_reputation": _join_export_value(sp.get("asn_takedown_reputation", {})),
            "silentpush_asn_ip_reputation_history": _join_export_value(sp.get("asn_ip_reputation_history", {})),
            "silentpush_asn_takedown_reputation_history": _join_export_value(sp.get("asn_takedown_reputation_history", {})),
            "silentpush_is_listed": sp.get("is_listed"),
            "silentpush_listed_txt": sp.get("listed_txt"),
            "silentpush_asn": sp.get("asn"),
            "silentpush_as_org": sp.get("as_org"),
            "silentpush_country": sp.get("country"),
            "silentpush_registrar": sp.get("registrar"),
            "silentpush_first_seen": sp.get("first_seen"),
            "silentpush_last_seen": sp.get("last_seen"),
            "silentpush_tags": _join_export_value(sp.get("tags", [])),
            "silentpush_nameservers": _join_export_value(sp.get("nameservers", [])),
            "silentpush_dns_values": _join_export_value(sp.get("dns_values", [])),
            "silentpush_available_endpoints": _join_export_value(sp.get("available_endpoints", [])),
            "silentpush_unavailable_endpoints": _join_export_value(sp.get("unavailable_endpoints", [])),
            "silentpush_enrichment": _join_export_value(sp.get("enrichment", {})),
            "silentpush_non_explore_available_apis": _join_export_value(sp.get("non_explore_available_apis", [])),
            "silentpush_non_explore_unavailable_attempts": _join_export_value(sp.get("non_explore_unavailable_attempts", [])),
            "silentpush_non_explore_api_results": _join_export_value(sp.get("non_explore_api_results", {})),
            "silentpush_threat_rank": sp.get("threat_rank"),
            "silentpush_live_scan_status": sp.get("live_scan_status"),
            "silentpush_context_graph_hits": sp.get("context_graph_hits"),
            "silentpush_bulletproof_hosting_likelihood": sp.get("bulletproof_hosting_likelihood"),
            "silentpush_bulletproof_hosting_reason": sp.get("bulletproof_hosting_reason"),
            "silentpush_bulletproof_hosting_flag": sp.get("bulletproof_hosting_flag"),
            "silentpush_padns_reverse_a_domains_count": sp.get("padns_reverse_a_domains_count"),
            "silentpush_padns_reverse_a_domains_sample": _join_export_value(sp.get("padns_reverse_a_domains_sample", [])),
            "silentpush_error": sp.get("error"),
        })

        row_copy.update({
            "spur_ip": spur.get("ip"),
            "spur_infrastructure": spur.get("infrastructure"),
            "spur_risks": _join_export_value(spur.get("risks", [])),
            "spur_services": _join_export_value(spur.get("services", [])),
            "spur_tunnel_count": spur.get("tunnel_count"),
            "spur_proxy_count": spur.get("proxy_count"),
            "spur_asn": spur.get("asn"),
            "spur_as_organization": spur.get("as_organization"),
            "spur_organization": spur.get("organization"),
            "spur_country": spur.get("country"),
            "spur_city": spur.get("city"),
            "spur_client_behaviors": _join_export_value(spur.get("client_behaviors", [])),
            "spur_balance_remaining": spur.get("balance_remaining"),
            "spur_result_dt": spur.get("result_dt"),
            "spur_concentration_country": spur.get("concentration_country"),
            "spur_concentration_state": spur.get("concentration_state"),
            "spur_concentration_city": spur.get("concentration_city"),
            "spur_concentration_density": spur.get("concentration_density"),
            "spur_concentration_skew": spur.get("concentration_skew"),
            "spur_concentration_geohash": spur.get("concentration_geohash"),
            "spur_client_count": spur.get("client_count"),
            "spur_client_countries": spur.get("client_countries"),
            "spur_client_spread": spur.get("client_spread"),
            "spur_client_types": _join_export_value(spur.get("client_types", [])),
            "spur_tunnel_types": _join_export_value(spur.get("tunnel_types", [])),
            "spur_tunnel_operators": _join_export_value(spur.get("tunnel_operators", [])),
            "spur_tunnel_tags": _join_export_value(spur.get("tunnel_tags", [])),
            "spur_error": spur.get("error"),
        })

        row_copy.update({
            "rl_endpoint": rl.get("endpoint"),
            "rl_verdict": rl.get("rl_verdict"),
            "rl_classification": rl.get("classification"),
            "rl_threat_level": rl.get("threat_level"),
            "rl_risk_score": rl.get("risk_score"),
            "rl_malicious_signals": rl.get("malicious_signals"),
            "rl_registrar": rl.get("registrar"),
            "rl_registrant": rl.get("registrant"),
            "rl_organization": rl.get("organization"),
            "rl_country": rl.get("country"),
            "rl_created": rl.get("created"),
            "rl_updated": rl.get("updated"),
            "rl_expires": rl.get("expires"),
            "rl_asn": rl.get("asn"),
            "rl_as_owner": rl.get("as_owner"),
            "rl_network": rl.get("network"),
            "rl_registry": rl.get("registry"),
            "rl_nameservers": _join_export_value(rl.get("nameservers", [])),
            "rl_third_party_malicious": rl.get("third_party_malicious"),
            "rl_third_party_suspicious": rl.get("third_party_suspicious"),
            "rl_third_party_clean": rl.get("third_party_clean"),
            "rl_third_party_total": rl.get("third_party_total"),
            "rl_third_party_detection_ratio": rl.get("third_party_detection_ratio"),
            "rl_highlights": _join_export_value(rl.get("highlights", {})),
            "rl_related_urls_count": rl.get("related_urls_count"),
            "rl_related_urls_sample": _join_export_value(rl.get("related_urls_sample", [])),
            "rl_resolutions_count": rl.get("resolutions_count"),
            "rl_resolutions_sample": _join_export_value(rl.get("resolutions_sample", [])),
            "rl_downloaded_files_count": rl.get("downloaded_files_count"),
            "rl_downloaded_files_sample": _join_export_value(rl.get("downloaded_files_sample", [])),
            "rl_error": rl.get("error"),
        })

        row_copy["sources_json"] = json.dumps(sources, separators=(",", ":"))
        flat_rows.append(row_copy)
    return flat_rows


def format_results(results, fmt: str) -> str:
    rows = results if isinstance(results, list) else [results]
    if fmt == "csv":
        buf = io.StringIO()
        writer = csv.DictWriter(buf, fieldnames=EXPORT_FIELDNAMES, extrasaction="ignore")
        writer.writeheader()

        for row_copy in _flatten_export_rows(rows):
            writer.writerow(row_copy)
        return buf.getvalue()
    return json.dumps(results, indent=2)


def _autosize_worksheet_columns(worksheet, max_width: int = 60) -> None:
    for column_cells in worksheet.columns:
        column = column_cells[0].column_letter
        max_len = 0
        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            if len(value) > max_len:
                max_len = len(value)
        worksheet.column_dimensions[column].width = min(max(10, max_len + 2), max_width)


def write_xlsx_results(results, output_path: str) -> None:
    try:
        from openpyxl import Workbook
    except ImportError:
        print("[ERROR] openpyxl is required to write XLSX files. Install it with: pip install openpyxl", file=sys.stderr)
        sys.exit(1)

    rows = results if isinstance(results, list) else [results]
    detailed_rows = _flatten_export_rows(rows)

    workbook = Workbook()

    summary_fields = [
        "ioc", "type", "verdict", "score", "time_taken", "reason", "reasons",
        "source_count", "source_errors",
    ]
    ws_summary = workbook.active
    ws_summary.title = "Summary"
    ws_summary.append(summary_fields)
    for row in detailed_rows:
        ws_summary.append([row.get(k, "") for k in summary_fields])
    ws_summary.freeze_panes = "A2"
    ws_summary.auto_filter.ref = ws_summary.dimensions
    _autosize_worksheet_columns(ws_summary)

    ws_detailed = workbook.create_sheet("Detailed")
    ws_detailed.append(EXPORT_FIELDNAMES)
    for row in detailed_rows:
        ws_detailed.append([row.get(k, "") for k in EXPORT_FIELDNAMES])
    ws_detailed.freeze_panes = "A2"
    ws_detailed.auto_filter.ref = ws_detailed.dimensions
    _autosize_worksheet_columns(ws_detailed)

    ws_raw = workbook.create_sheet("RawJSON")
    ws_raw.append(["ioc", "json"])
    for row in rows:
        ws_raw.append([row.get("ioc", ""), json.dumps(row, ensure_ascii=False)])
    ws_raw.freeze_panes = "A2"
    ws_raw.auto_filter.ref = ws_raw.dimensions
    _autosize_worksheet_columns(ws_raw, max_width=120)

    workbook.save(output_path)


# ----------------------------
# CLI
# ----------------------------
def main():
    check_keys()

    parser = argparse.ArgumentParser(description="IOC threat lookup tool")
    parser.add_argument("ioc", nargs="?", help="Single IOC to analyze (IP, domain, hash, or URL)")
    parser.add_argument("--file",   metavar="PATH",  help="Text file with one IOC per line")
    parser.add_argument("--output", metavar="PATH",  help="Write results to this file instead of stdout")
    parser.add_argument("--format", choices=["json", "csv", "xlsx"], default="json", help="Output format (default: json)")
    parser.add_argument("--delay",  type=float, default=1.0, metavar="SECS",
                        help="Global minimum seconds between IOC starts in bulk mode (default: 1.0)")
    parser.add_argument("--workers", type=int, default=2, metavar="N",
                        help="Concurrent IOC workers in bulk mode (default: 2)")
    parser.add_argument("--jitter", type=float, default=0.25, metavar="SECS",
                        help="Random extra delay added per IOC start to avoid burst patterns (default: 0.25)")
    parser.add_argument("--batch-size", type=int, default=150, metavar="N",
                        help="Maximum IOCs processed per batch before cooldown (default: 150)")
    parser.add_argument("--batch-cooldown", type=float, default=30.0, metavar="SECS",
                        help="Pause between batches in bulk mode to avoid API bursts (default: 30)")
    parser.add_argument("--max-iocs", type=int, default=0, metavar="N",
                        help="Optional cap on number of unique IOCs to process (0 = no cap)")

    args = parser.parse_args()

    if not args.ioc and not args.file:
        parser.error("provide either an IOC argument or --file PATH")
    if args.ioc and args.file:
        parser.error("provide either an IOC argument or --file PATH, not both")

    if args.file:
        ext = Path(args.file).suffix.lower()
        if ext in (".xlsx", ".xls"):
            iocs = _read_iocs_excel(args.file)
        else:
            try:
                with open(args.file, "r", encoding="utf-8-sig") as f:
                    iocs = [l.strip() for l in f if l.strip() and not l.strip().startswith("#")]
            except OSError as e:
                print(f"[ERROR] Cannot read file: {e}", file=sys.stderr)
                sys.exit(1)

        iocs = _prepare_iocs(iocs, args.max_iocs)
        if not iocs:
            print("[ERROR] No valid IOCs found after filtering/deduplication.", file=sys.stderr)
            sys.exit(1)

        batch_size = max(1, args.batch_size)
        batch_cooldown = max(0.0, args.batch_cooldown)

        if len(iocs) > batch_size:
            output = []
            total_batches = (len(iocs) + batch_size - 1) // batch_size
            for batch_idx in range(total_batches):
                start = batch_idx * batch_size
                end = min(len(iocs), start + batch_size)
                batch = iocs[start:end]
                print(
                    f"[INFO] Batch {batch_idx+1}/{total_batches}: processing {len(batch)} IOC(s)",
                    file=sys.stderr,
                )
                output.extend(asyncio.run(analyze_bulk(batch, args.workers, args.delay, args.jitter)))

                if batch_idx < total_batches - 1 and batch_cooldown > 0:
                    print(
                        f"[INFO] Cooling down for {batch_cooldown:.1f}s before next batch...",
                        file=sys.stderr,
                    )
                    time.sleep(batch_cooldown)
        else:
            output = asyncio.run(analyze_bulk(iocs, args.workers, args.delay, args.jitter))
    else:
        output = asyncio.run(analyze(args.ioc))

    if args.format == "xlsx":
        if not args.output:
            print("[ERROR] XLSX output requires --output PATH", file=sys.stderr)
            sys.exit(1)
        try:
            write_xlsx_results(output, args.output)
            print(f"[INFO] Results written to {args.output}", file=sys.stderr)
        except OSError as e:
            print(f"[ERROR] Cannot write output file: {e}", file=sys.stderr)
            sys.exit(1)
    else:
        formatted = format_results(output, args.format)
        if args.output:
            try:
                with open(args.output, "w", encoding="utf-8") as f:
                    f.write(formatted)
                print(f"[INFO] Results written to {args.output}", file=sys.stderr)
            except OSError as e:
                print(f"[ERROR] Cannot write output file: {e}", file=sys.stderr)
                sys.exit(1)
        else:
            print(formatted)


if __name__ == "__main__":
    main()